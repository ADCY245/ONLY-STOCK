import json
import os
import re
import math
from datetime import datetime, timezone as datetime_timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import xlrd
import xlwt
from bson import ObjectId
from flask import Flask, jsonify, request, send_file, send_from_directory, session
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from pymongo.errors import DuplicateKeyError, PyMongoError
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from .db import get_database, get_inventory_collection, get_stock_logs_collection, get_users_collection
except ImportError:
    from db import get_database, get_inventory_collection, get_stock_logs_collection, get_users_collection


app = Flask(__name__)
CORS(app, supports_credentials=True)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "only-stock-dev-secret")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"
ALLOWED_SIGNUP_DOMAIN = "@chemo.in"
SUPERADMIN_EMAIL = "athulnair3096@gmail.com"
ALLOWED_ROLES = {"user", "admin", "workshop"}
READ_ONLY_ROLES = {"user"}
WRITE_ROLES = {"admin", "workshop"}
LOG_ACCESS_ROLES = {"admin"}
LOCATION_ID_PATTERN = re.compile(r"^WH-A-[A-Z0-9]{2,8}-L[1-9][0-9]*-[A-Z]$")

REQUIRED_EXCEL_COLUMNS = [
    "Category",
    "Brand",
    "Type",
    "Width",
    "Length",
    "Thickness",
    "Quantity",
    "Unit",
]
EXCEL_COLUMNS = [
    "Category",
    "Brand",
    "Type",
    "Blanket Name",
    "Storage Type",
    "Batch/Roll No",
    "Width",
    "Nominal Width",
    "Actual Width",
    "Width Unit",
    "Length",
    "Length Unit",
    "Size",
    "Thickness",
    "Thickness Unit",
    "Thickness (Micron)",
    "Rolls",
    "Number of Rolls",
    "No. of Rolls",
    "Number of Sheets",
    "No. of Sheets",
    "Roll No",
    "Batch No",
    "Print Type",
    "Area per Roll",
    "Area per Sheet",
    "Quantity",
    "Unit",
    "Product",
    "Pack Size",
    "Container Type",
    "Containers/Box",
    "Boxes",
    "Loose Units",
    "Containers",
    "Total Sheets",
]
INVENTORY_ADJUSTMENT_COLUMNS = [
    "Reference#",
    "InventoryAdjustment#",
    "Date",
    "Status",
    "InventoryAdjustment ID",
    "Adjustment Type",
    "Reason",
    "Item Name",
    "Item ID",
    "Batch Number",
    "Batch Reference#",
    "Manufacturer Batch#",
    "Warehouse Name",
    "Account",
    "Inventory Account",
    "Quantity Adjusted",
    "Cost Price",
    "Value Adjusted",
    "Usage unit",
]
DEFAULT_LOW_STOCK_THRESHOLD = 5
DEFAULT_ADJUSTMENT_ITEM_NAME = "Image - Print Master BL - 1070 mm - 1.95mm"
FIXED_ADJUSTMENT_REASON = "Stock Update based on 1.04.26"
DEFAULT_ADJUSTMENT_WAREHOUSE = "Main Location"
IST_TIMEZONE = ZoneInfo("Asia/Kolkata")
ROLL_PAPER_CATEGORY = "Calibrated Underpacking Paper"
ROLL_PAPER_STOCK_UNIT = "m²"
M3Z_CATEGORY_DISPLAY = "05 - Calibrated Underpacking Paper"
M3Z_ALLOWED_THICKNESS_MICRONS = {100, 150, 200, 250, 300, 400, 500}
M3Z_ROLL_STORAGE_TYPE = "roll"
M3Z_CUT_PIECE_STORAGE_TYPE = "cut_piece"
M3Z_SHEET_STOCK_UNIT = "sheets"
RUBBER_BLANKET_CATEGORY = "Rubber Blankets"
RUBBER_BLANKET_STOCK_UNIT = "m²"
ROLL_WIDTH_UNITS = {"mm", "m", "inch"}
ROLL_LENGTH_UNITS = {"m", "mm", "inch"}
ROLL_THICKNESS_UNITS = {"mm", "micron"}

# Rubber Blanket catalog source of truth. The same JSON file is fetched by the
# browser, so names and validation rules do not need to be edited in two files.
RUBBER_BLANKET_CONFIG_PATH = FRONTEND_DIR / "data" / "inventory-config.json"


def load_rubber_blanket_rules():
    def config_text(value):
        return str(value).strip() if value is not None else ""

    try:
        with RUBBER_BLANKET_CONFIG_PATH.open("r", encoding="utf-8") as config_file:
            config = json.load(config_file)
    except (OSError, ValueError) as error:
        raise RuntimeError(f"Unable to load inventory config: {RUBBER_BLANKET_CONFIG_PATH}") from error

    entries = config.get("rubber_blankets") if isinstance(config, dict) else None
    if not isinstance(entries, list) or not entries:
        raise RuntimeError("inventory-config.json must contain a non-empty rubber_blankets list")

    rules = {}
    for entry in entries:
        if not isinstance(entry, dict) or not config_text(entry.get("name")):
            raise RuntimeError("Each rubber blanket config entry must have a name")
        name = config_text(entry["name"])
        widths = entry.get("widths", [])
        thickness_options = entry.get("thickness_options", [])
        if not isinstance(widths, list) or not widths or not isinstance(thickness_options, list) or not thickness_options:
            raise RuntimeError(f"Invalid dimensions or thickness options for {name}")
        rules[name] = {
            "aliases": [config_text(value) for value in entry.get("aliases", []) if config_text(value)],
            "thickness_mode": config_text(entry.get("thickness_mode")) or "select",
            "thickness": entry.get("thickness"),
            "thickness_options": [float(value) for value in thickness_options],
            "widths": [tuple(float(value) for value in pair) for pair in widths if isinstance(pair, (list, tuple)) and len(pair) == 2],
            "print_types": [config_text(value).upper() for value in entry.get("print_types", []) if config_text(value)],
        }
    if len(rules) != len(entries):
        raise RuntimeError("Rubber blanket names must be unique in inventory-config.json")
    return rules


RUBBER_BLANKET_RULES = load_rubber_blanket_rules()
THICKNESS_REQUIRED_CATEGORIES = {
    "Rubber Blankets",
    "Metalback Blankets",
    "Underlay Blanket",
    "Calibrated Underpacking Paper",
    "Calibrated Underpacking Film",
    "Creasing Matrix",
    "Cutting Rules",
    "Creasing Rules",
    "Litho Perforation Rules",
    "CTCP Plates",
}
DIMENSIONAL_CATEGORIES = THICKNESS_REQUIRED_CATEGORIES | {
    "Blanket Barring",
    "Cutting String",
    "Ejection Rubber",
    "Strip Plate",
    "Anti Marking Film",
    "Ink Duct Foil",
    "Productive Foil",
    "Presspahn Sheets",
    "Auto Wash Cloth",
    "ICP Paper",
    "Dampening Hose",
    "Tesamol Tape",
}

NO_BRAND_TYPE_CATEGORIES = {
    "Creasing Matrix",
    "CTCP Plates",
    ROLL_PAPER_CATEGORY,
}

RULE_UNIT_LINKED_CATEGORIES = {
    "Cutting Rules",
    "Creasing Rules",
    "Litho Perforation Rules",
}

CHEMICAL_CATEGORIES = {
    "Washing Solutions",
    "Fountain Solutions",
    "Plate Care Products",
    "Roller Care Products",
    "Blanket Maintenance Products",
}

BLANKET_BATCH_ROLL_CATEGORIES = {
    "Rubber Blankets",
    "Metalback Blankets",
}

ALLOWED_CATEGORIES = {
    "Rubber Blankets",
    "Metalback Blankets",
    "Underlay Blanket",
    "Blanket Barring",
    "Calibrated Underpacking Paper",
    "Calibrated Underpacking Film",
    "Creasing Matrix",
    "Cutting Rules",
    "Creasing Rules",
    "Litho Perforation Rules",
    "Cutting String",
    "Ejection Rubber",
    "Strip Plate",
    "Anti Marking Film",
    "Ink Duct Foil",
    "Productive Foil",
    "Presspahn Sheets",
    "Washing Solutions",
    "Fountain Solutions",
    "Plate Care Products",
    "Roller Care Products",
    "Blanket Maintenance Products",
    "Auto Wash Cloth",
    "ICP Paper",
    "Spray Powder",
    "Sponges",
    "Dampening Hose",
    "Tesamol Tape",
    "CTCP Plates",
}


def normalize_category_name(value):
    category = clean_text(value)
    if not category:
        return None
    prefixed_match = re.match(r"^\d+\s*-\s*(.+)$", category)
    if prefixed_match:
        category = prefixed_match.group(1).strip()
    for allowed_category in ALLOWED_CATEGORIES:
        if allowed_category.casefold() == category.casefold():
            return allowed_category
    return category

CREASING_MATRIX_SIZES = {
    "9": [
        "0.3 X 1.0", "0.3 X 1.1", "0.3 X 1.2", "0.3 X 1.3", "0.3 X 1.5",
        "0.4 X 1.0", "0.4 X 1.1", "0.4 X 1.2", "0.4 X 1.3", "0.4 X 1.4",
        "0.4 X 1.5", "0.4 X 1.6", "0.4 X 1.7", "0.5 X 1.2", "0.5 X 1.3",
        "0.5 X 1.4", "0.5 X 1.5", "0.5 X 1.6", "0.5 X 1.7", "0.5 X 1.8",
        "0.5 X 1.9", "0.6 X 1.1", "0.6 X 1.5", "0.6 X 1.6", "0.6 X 1.7",
        "0.6 X 1.9", "0.6 X 2.1", "0.6 X 2.3", "0.6 X 2.5",
    ],
    "11": [
        "0.7 X 1.1", "0.7 X 1.5", "0.7 X 2.1", "0.7 X 2.3", "0.7 X 2.5",
        "0.7 X 2.7", "0.8 X 2.1", "0.8 X 2.3", "0.8 X 2.5", "0.8 X 2.7",
        "0.8 X 3.0",
    ],
    "13": ["1.0 X 3.0", "1.0 X 3.2", "1.0 X 3.5", "1.0 X 4.0", "1.4 X 5.0"],
}

CTCP_PLATE_SIZES = {
    "0.30": ["650 X 550", "730 X 600", "620 X 482"],
    "0.20": ["520 X 400"],
}

SPECIALIZED_CHEMICAL_PRODUCTS = {
    "Chem R-ol": {
        "category": "Washing Solutions",
        "display_format": "Chem R-ol (5L Pack)",
        "pack_size": 5,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [4],
    },
    "FS Clean": {
        "category": "Fountain Solutions",
        "display_format": "FS Clean (5L Pack)",
        "pack_size": 5,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [4],
    },
    "Anilox Clean": {
        "category": "Roller Care Products",
        "display_format": "Anilox Clean 5L",
        "pack_size": 5,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [4],
    },
    "Roll-o-clean": {
        "category": "Roller Care Products",
        "display_format": "Roll-o-clean 1kg",
        "pack_size": 1,
        "unit": "kg",
        "container_type": "bottle",
        "containers_per_box": [12, 15, 18],
    },
    "MT-R Clean": {
        "category": "Roller Care Products",
        "display_format": "MT-R Clean 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [12],
    },
    "Blanket Clean": {
        "category": "Blanket Maintenance Products",
        "display_format": "Blanket Clean 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [5],
    },
    "Blanket Clean UV": {
        "category": "Blanket Maintenance Products",
        "display_format": "Blanket Clean UV 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [5],
    },
    "ALU Clean": {
        "category": "Plate Care Products",
        "display_format": "ALU Clean 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [12],
    },
    "ALU Clean UV": {
        "category": "Plate Care Products",
        "display_format": "ALU Clean UV 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [12],
    },
    "Calx De Glazer": {
        "category": "Roller Care Products",
        "display_format": "Calx De Glazer 1L",
        "pack_size": 1,
        "unit": "ltr",
        "container_type": "bottle",
        "containers_per_box": [12],
    },
}

EXCEL_README_SHEET = "README - Instructions"
EXCEL_LISTS_SHEET = "_Lists"
EXCEL_LEGACY_SHEET = "Legacy_Inventory"
EXCEL_HEADER_ROW = 4
EXCEL_DATA_START_ROW = EXCEL_HEADER_ROW + 1
EXCEL_MAX_INPUT_ROW = 504
EXCEL_CATEGORY_SHEETS = (
    {
        "key": "rubber_blankets",
        "sheet_name": "01_Rubber_Blankets",
        "label": "Rubber Blankets",
        "accent": "15803D",
        "headers": [
            "Blanket Name", "Storage Type", "Thickness", "Thickness Unit", "Print Type",
            "Nominal Width", "Actual Width", "Width Unit", "Length", "Length Unit",
            "Number of Rolls", "Number of Sheets", "Area per Roll", "Area per Sheet", "Quantity", "Unit",
        ],
        "required_headers": [
            "Blanket Name", "Storage Type", "Thickness", "Thickness Unit", "Print Type",
            "Nominal Width", "Actual Width", "Width Unit", "Length", "Length Unit",
            "Number of Rolls", "Number of Sheets",
        ],
        "derived_headers": ["Area per Roll", "Area per Sheet", "Quantity", "Unit"],
        "instruction": (
            "Required fields are red. Choose Roll or Cut Piece; fill Number of Rolls or Number of Sheets accordingly. "
            "Area and Quantity are recalculated by Only Stock."
        ),
    },
    {
        "key": "calibrated_underpacking_paper",
        # Excel limits worksheet titles to 31 characters. The requested long title is
        # accepted as an import alias below, while generated workbooks use this safe title.
        "sheet_name": "02_Calibrated_Underpacking",
        "label": "Calibrated Underpacking Paper",
        "accent": "2563EB",
        "headers": [
            "Storage Type", "Thickness (Micron)", "Width", "Width Unit", "Length", "Length Unit",
            "Number of Rolls", "Number of Sheets", "Area per Roll", "Area per Sheet", "Quantity", "Unit",
        ],
        "required_headers": [
            "Storage Type", "Thickness (Micron)", "Width", "Width Unit", "Length", "Length Unit",
            "Number of Rolls", "Number of Sheets",
        ],
        "derived_headers": ["Area per Roll", "Area per Sheet", "Quantity", "Unit"],
        "instruction": (
            "Required fields are red. Thickness is entered in micron. Roll stock is normalized to Sq.m; "
            "Cut Piece stock is stored in Sheets."
        ),
    },
    {
        "key": "creasing_matrix",
        "sheet_name": "03_Creasing_Matrix",
        "label": "Creasing Matrix",
        "accent": "EA580C",
        "headers": ["Thickness", "Size", "Boxes", "Loose Packets", "Quantity (Pkt)", "Unit"],
        "required_headers": ["Thickness", "Size", "Boxes", "Loose Packets", "Quantity (Pkt)"],
        "derived_headers": ["Quantity (Pkt)", "Unit"],
        "instruction": (
            "Required fields are red. Enter Boxes and Loose Packets, or Quantity (Pkt). "
            "Only Stock validates the thickness/size pair and stores stock in packets."
        ),
    },
    {
        "key": "chemical",
        "sheet_name": "04_Chemical",
        "label": "Chemical",
        "accent": "7C3AED",
        "headers": [
            "Product", "Pack Size", "Container Type", "Containers per Box", "Boxes", "Loose Containers",
            "Total Containers", "Total Quantity", "Unit",
        ],
        "required_headers": [
            "Product", "Pack Size", "Container Type", "Containers per Box", "Boxes", "Loose Containers",
        ],
        "derived_headers": ["Total Containers", "Total Quantity", "Unit"],
        "instruction": (
            "Required fields are red. Choose a product and packaging configuration. "
            "Only Stock recalculates total containers and normalized litre/kg quantity."
        ),
    },
    {
        "key": "ctcp_plates",
        "sheet_name": "05_CTCP_Plates",
        "label": "CTCP Plates",
        "accent": "CA8A04",
        "headers": ["Thickness", "Size", "Boxes", "Sheets per Box", "Total Sheets", "Quantity", "Unit"],
        "required_headers": ["Thickness", "Size", "Boxes"],
        "derived_headers": ["Sheets per Box", "Total Sheets", "Quantity", "Unit"],
        "instruction": (
            "Required fields are red. Only Stock validates the thickness/size pair. "
            "One box contains 50 sheets; inventory remains stored in Boxes."
        ),
    },
)
EXCEL_SHEET_BY_NAME = {definition["sheet_name"]: definition for definition in EXCEL_CATEGORY_SHEETS}
EXCEL_SHEET_BY_KEY = {definition["key"]: definition for definition in EXCEL_CATEGORY_SHEETS}
EXCEL_SHEET_BY_NAME["02_Calibrated_Underpacking_Paper"] = EXCEL_SHEET_BY_KEY["calibrated_underpacking_paper"]


def now_ist():
    return datetime.now(IST_TIMEZONE)


def serialize_datetime_ist(value):
    if not value:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=datetime_timezone.utc)
    return value.astimezone(IST_TIMEZONE).isoformat()


def clean_text(value):
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if not isinstance(value, str):
        return None
    cleaned = value.strip()
    return cleaned or None


def normalize_email(value):
    cleaned = clean_text(value)
    return cleaned.lower() if cleaned else None


def serialize_user(user):
    return {
        "id": str(user["_id"]),
        "email": user["email"],
        "role": user["role"],
        "is_superadmin": user["email"] == SUPERADMIN_EMAIL,
        "created_at": serialize_datetime_ist(user.get("created_at")),
        "updated_at": serialize_datetime_ist(user.get("updated_at")),
    }


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    try:
        return get_users_collection().find_one({"_id": ObjectId(user_id)})
    except Exception:
        session.clear()
        return None


def require_auth():
    user = get_current_user()
    if not user:
        return None, (jsonify({"error": "Authentication required"}), 401)
    return user, None


def require_role(*roles):
    user, error_response = require_auth()
    if error_response:
        return None, error_response
    if user["role"] not in roles:
        return None, (jsonify({"error": "You do not have permission for this action"}), 403)
    return user, None


def create_superadmin_if_missing(password):
    users_collection = get_users_collection()
    user = users_collection.find_one({"email": SUPERADMIN_EMAIL})
    if user:
        return user, False

    now = now_ist()
    user = {
        "email": SUPERADMIN_EMAIL,
        "password_hash": generate_password_hash(password.strip()),
        "role": "admin",
        "created_at": now,
        "updated_at": now,
    }
    result = users_collection.insert_one(user)
    user["_id"] = result.inserted_id
    return user, True


def validate_signup_email(email):
    if email == SUPERADMIN_EMAIL:
        return None
    if not email.endswith(ALLOWED_SIGNUP_DOMAIN):
        return "Signup is not permitted for this email"
    return None


def validate_password(password):
    if not isinstance(password, str) or len(password.strip()) < 6:
        return "Password must be at least 6 characters"
    return None


def parse_integer(value, field_name, allow_negative=False):
    if isinstance(value, bool):
        return None, f"{field_name} must be an integer"

    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float) and value.is_integer():
        parsed = int(value)
    elif isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None, f"{field_name} is required"
        if re.fullmatch(r"-?\d+", stripped) is None:
            return None, f"{field_name} must be an integer"
        parsed = int(stripped)
    else:
        return None, f"{field_name} must be an integer"

    if not allow_negative and parsed < 0:
        return None, f"{field_name} must be a non-negative integer"

    return parsed, None


def parse_number(value, field_name, allow_negative=False):
    if isinstance(value, bool):
        return None, f"{field_name} must be a number"

    if isinstance(value, (int, float)):
        parsed = float(value)
    elif isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None, f"{field_name} is required"
        try:
            parsed = float(stripped)
        except ValueError:
            return None, f"{field_name} must be a number"
    else:
        return None, f"{field_name} must be a number"

    if not math.isfinite(parsed):
        return None, f"{field_name} must be a finite number"

    if not allow_negative and parsed < 0:
        return None, f"{field_name} must be a non-negative number"

    return parsed, None


def parse_dimension_number(value):
    text = str(value or "").strip()
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def calculate_roll_area_sqm(width, length):
    width_mm = parse_dimension_number(width)
    length_mtr = parse_dimension_number(length)
    if width_mm is None or length_mtr is None:
        return None
    return round((width_mm / 1000) * length_mtr, 4)


def format_numeric_text(value, max_decimals=6):
    formatted = f"{float(value):.{max_decimals}f}".rstrip("0").rstrip(".")
    return formatted or "0"


def normalize_roll_width_unit(value):
    unit = (clean_text(value) or "").lower()
    return unit if unit in ROLL_WIDTH_UNITS else None


def normalize_roll_length_unit(value):
    unit = (clean_text(value) or "").lower()
    return unit if unit in ROLL_LENGTH_UNITS else None


def normalize_roll_thickness_unit(value):
    unit = (clean_text(value) or "").lower()
    if unit in {"mm"}:
        return "mm"
    if unit in {"micron", "microns", "um", "µm", "μm"}:
        return "micron"
    return None


def normalize_roll_stock_unit(value):
    unit = (clean_text(value) or "").lower().replace(" ", "")
    if unit in {"m²", "m2", "sqm", "sq.m", "sq.mtr", "square.mtr", "squaremeter", "squaremeters"}:
        return ROLL_PAPER_STOCK_UNIT
    return None


def normalize_sheet_stock_unit(value):
    unit = (clean_text(value) or "").lower().replace(" ", "")
    if unit in {"sheet", "sheets", "pcs", "pieces", "cutpiece", "cutpieces"}:
        return M3Z_SHEET_STOCK_UNIT
    return None


def parse_positive_number(value, field_name):
    parsed, error = parse_number(value, field_name)
    if error:
        return None, error
    if parsed <= 0:
        return None, f"{field_name} must be greater than 0"
    return parsed, None


def parse_positive_integer(value, field_name):
    parsed, error = parse_integer(value, field_name)
    if error:
        return None, error
    if parsed <= 0:
        return None, f"{field_name} must be greater than 0"
    return parsed, None


def parse_roll_thickness(data):
    raw_value = parse_optional_text(data.get("thickness"))
    if not raw_value:
        return None, None, "thickness is required", None

    supplied_unit = clean_text(data.get("thickness_unit"))
    thickness_unit = normalize_roll_thickness_unit(supplied_unit)
    if supplied_unit and not thickness_unit:
        return None, None, "thickness unit must be mm or micron", None
    numeric_value = raw_value
    if not thickness_unit:
        embedded_unit = re.search(r"\s*(mm|microns?|um|µm|μm)\s*$", raw_value, flags=re.IGNORECASE)
        if embedded_unit:
            thickness_unit = normalize_roll_thickness_unit(embedded_unit.group(1))
            numeric_value = raw_value[:embedded_unit.start()].strip()
        else:
            thickness_unit = "mm"

    thickness, error = parse_positive_number(numeric_value, "thickness")
    if error:
        return None, None, error, None
    thickness_micron = thickness * 1000 if thickness_unit == "mm" else thickness
    return thickness, thickness_unit, None, thickness_micron


def parse_m3z_thickness(data):
    raw_micron = data.get("thickness_micron")
    if parse_optional_text(raw_micron) is not None:
        thickness_micron, error = parse_positive_number(raw_micron, "thickness")
        if error:
            return None, error
        supplied_unit = clean_text(data.get("thickness_unit"))
        if supplied_unit:
            normalized_unit = normalize_roll_thickness_unit(supplied_unit)
            if not normalized_unit:
                return None, "thickness unit must be mm or micron"
    else:
        thickness_result = parse_roll_thickness(data)
        if thickness_result[2]:
            return None, thickness_result[2]
        _, _, _, thickness_micron = thickness_result
    if not math.isclose(thickness_micron, round(thickness_micron), rel_tol=0, abs_tol=0.000001):
        return None, "thickness must be one of 100, 150, 200, 250, 300, 400, or 500 micron"
    thickness_micron = int(round(thickness_micron))
    if thickness_micron not in M3Z_ALLOWED_THICKNESS_MICRONS:
        return None, "thickness must be one of 100, 150, 200, 250, 300, 400, or 500 micron"
    return thickness_micron, None


def normalize_m3z_storage_type(value):
    storage_type = re.sub(r"[\s-]+", "_", (clean_text(value) or "").lower())
    if storage_type in {"roll", "rolls", "as_rolls", "as_roll"}:
        return M3Z_ROLL_STORAGE_TYPE
    if storage_type in {"cut_piece", "cut_pieces", "cutpiece", "cutpieces", "sheets", "sheet"}:
        return M3Z_CUT_PIECE_STORAGE_TYPE
    return None


def build_roll_paper_fields(data, require_rolls=True):
    raw_storage_type = clean_text(data.get("storage_type"))
    storage_type = normalize_m3z_storage_type(raw_storage_type) if raw_storage_type else M3Z_ROLL_STORAGE_TYPE
    if raw_storage_type and not storage_type:
        return None, "storage type must be roll or cut_piece"
    width, width_error = parse_positive_number(data.get("width"), "width")
    if width_error:
        return None, width_error
    width_unit = normalize_roll_width_unit(data.get("width_unit"))
    if not width_unit:
        return None, "width unit must be mm, m, or inch"

    raw_length = data.get("length")
    if parse_optional_text(raw_length) is None:
        raw_length = data.get("height")
    length, length_error = parse_positive_number(raw_length, "length")
    if length_error:
        return None, length_error
    length_unit = normalize_roll_length_unit(data.get("length_unit") or data.get("height_unit"))
    if not length_unit:
        return None, "length unit must be m, mm, or inch"

    thickness_micron, thickness_error = parse_m3z_thickness(data)
    if thickness_error:
        return None, thickness_error
    thickness = format_numeric_text(thickness_micron / 1000, 6)
    supplied_thickness_unit = normalize_roll_thickness_unit(data.get("thickness_unit"))
    if supplied_thickness_unit:
        thickness_unit = supplied_thickness_unit
    elif parse_optional_text(data.get("thickness_micron")) is not None:
        thickness_unit = "micron"
    elif re.search(r"\s*(mm|microns?|um|µm|μm)\s*$", parse_optional_text(data.get("thickness")) or "", flags=re.IGNORECASE):
        thickness_unit = "micron" if re.search(r"microns?|um|µm|μm", parse_optional_text(data.get("thickness")) or "", flags=re.IGNORECASE) else "mm"
    else:
        thickness_unit = "mm"

    width_meters = width * {"mm": 0.001, "m": 1, "inch": 0.0254}[width_unit]
    length_meters = length * {"m": 1, "mm": 0.001, "inch": 0.0254}[length_unit]
    area_per_roll_sqm = width_meters * length_meters

    if storage_type == M3Z_ROLL_STORAGE_TYPE:
        raw_rolls = data.get("number_of_rolls")
        if parse_optional_text(raw_rolls) is None:
            raw_rolls = data.get("rolls")
        if parse_optional_text(raw_rolls) is None and not require_rolls:
            number_of_rolls = None
        else:
            number_of_rolls, rolls_error = parse_positive_integer(raw_rolls, "number of rolls")
            if rolls_error:
                return None, rolls_error
        number_of_sheets = None
        total_quantity = area_per_roll_sqm * (number_of_rolls or 0)
        expected_unit = ROLL_PAPER_STOCK_UNIT
    else:
        raw_sheets = data.get("number_of_sheets")
        if parse_optional_text(raw_sheets) is None:
            raw_sheets = data.get("sheets")
        if parse_optional_text(raw_sheets) is None:
            if require_rolls:
                return None, "number of sheets is required"
            number_of_sheets = None
        else:
            number_of_sheets, sheets_error = parse_positive_integer(raw_sheets, "number of sheets")
            if sheets_error:
                return None, sheets_error
        number_of_rolls = None
        total_quantity = number_of_sheets or 0
        expected_unit = M3Z_SHEET_STOCK_UNIT

    supplied_unit = clean_text(data.get("unit"))
    if storage_type == M3Z_ROLL_STORAGE_TYPE:
        if supplied_unit and not normalize_roll_stock_unit(supplied_unit):
            return None, "unit must be m² for Calibrated Underpacking Paper rolls"
    elif supplied_unit and not normalize_sheet_stock_unit(supplied_unit):
        return None, "unit must be sheets for Calibrated Underpacking Paper cut pieces"

    return {
        "storage_type": storage_type,
        "width": format_numeric_text(width),
        "width_unit": width_unit,
        "height": format_numeric_text(length),
        "length": format_numeric_text(length),
        "length_unit": length_unit,
        "thickness": thickness,
        "thickness_unit": thickness_unit,
        "thickness_micron": thickness_micron,
        "number_of_rolls": number_of_rolls,
        "number_of_sheets": number_of_sheets,
        "width_meters": round(width_meters, 9),
        "length_meters": round(length_meters, 9),
        "area_per_roll_sqm": round(area_per_roll_sqm, 6) if storage_type == M3Z_ROLL_STORAGE_TYPE else None,
        "area_per_sheet_sqm": round(area_per_roll_sqm, 6),
        "total_area_sqm": round(total_quantity, 6) if storage_type == M3Z_ROLL_STORAGE_TYPE else None,
        "total_quantity": round(total_quantity, 6),
        "unit": expected_unit,
    }, None


def normalize_blanket_name(value):
    return re.sub(r"\s+", " ", clean_text(value) or "").strip().casefold()


def find_rubber_blanket_rule(value):
    requested = normalize_blanket_name(value)
    if not requested:
        return None, None
    for name, rule in RUBBER_BLANKET_RULES.items():
        candidates = [name, *rule.get("aliases", [])]
        if any(normalize_blanket_name(candidate) == requested for candidate in candidates):
            return name, rule
    return None, None


def normalize_blanket_print_type(value):
    normalized = re.sub(r"\s+", " ", clean_text(value) or "").strip().upper()
    aliases = {
        "P": "P",
        "PRINTED": "P",
        "PRINTED (P)": "P",
        "W/O": "W/O",
        "WO": "W/O",
        "WITHOUT PRINT": "W/O",
        "WITHOUT PRINT (W/O)": "W/O",
    }
    return aliases.get(normalized)


def get_rubber_blanket_widths(rule, thickness):
    widths_by_thickness = rule.get("widths_by_thickness")
    if isinstance(widths_by_thickness, dict):
        thickness_key = format_numeric_text(thickness, 2)
        return widths_by_thickness.get(thickness_key, [])
    return rule.get("widths", [])


def build_rubber_blanket_fields(data, require_rolls=True):
    requested_name = data.get("blanket_name") or data.get("brand")
    blanket_name, rule = find_rubber_blanket_rule(requested_name)
    if not rule:
        return None, "invalid blanket name"

    raw_storage_type = clean_text(data.get("storage_type"))
    storage_type = normalize_m3z_storage_type(raw_storage_type) if raw_storage_type else M3Z_ROLL_STORAGE_TYPE
    if raw_storage_type and not storage_type:
        return None, "storage type must be roll or cut_piece"

    supplied_thickness_unit = clean_text(data.get("thickness_unit"))
    if supplied_thickness_unit and supplied_thickness_unit.lower() != "mm":
        return None, "Rubber Blanket thickness unit must be mm"

    raw_thickness = parse_optional_text(data.get("thickness"))
    if raw_thickness:
        thickness_text = re.sub(r"\s*mm\s*$", "", raw_thickness, flags=re.IGNORECASE).strip()
        thickness, thickness_error = parse_positive_number(thickness_text, "thickness")
        if thickness_error:
            return None, thickness_error
    elif rule.get("thickness_mode") == "fixed":
        thickness = float(rule["thickness"])
    else:
        return None, "thickness is required for this blanket"

    allowed_thicknesses = [float(value) for value in rule.get("thickness_options", [])]
    if not any(math.isclose(thickness, allowed, rel_tol=0, abs_tol=0.000001) for allowed in allowed_thicknesses):
        allowed_text = ", ".join(format_numeric_text(value, 2) for value in allowed_thicknesses)
        return None, f"thickness for {blanket_name} must be one of: {allowed_text} mm"
    if rule.get("thickness_mode") == "fixed" and not math.isclose(
        thickness, float(rule["thickness"]), rel_tol=0, abs_tol=0.000001
    ):
        return None, f"thickness for {blanket_name} must be {format_numeric_text(rule['thickness'], 2)} mm"

    width_unit = normalize_roll_width_unit(data.get("width_unit") or "mm")
    if not width_unit:
        return None, "Rubber Blanket width unit must be mm, m, or inch"
    width_scale_to_mm = {"mm": 1, "m": 1000, "inch": 25.4}[width_unit]

    raw_nominal_width = data.get("nominal_width")
    if parse_optional_text(raw_nominal_width) is None:
        raw_nominal_width = data.get("width")
    nominal_width, nominal_error = parse_positive_number(raw_nominal_width, "nominal width")
    if nominal_error:
        return None, nominal_error
    nominal_width_mm = nominal_width * width_scale_to_mm

    width_candidates = [
        (float(nominal), float(actual))
        for nominal, actual in get_rubber_blanket_widths(rule, thickness)
        if math.isclose(nominal_width_mm, float(nominal), rel_tol=0, abs_tol=0.000001)
    ]
    raw_actual_width = data.get("actual_width")
    if parse_optional_text(raw_actual_width) is not None:
        supplied_actual_width, actual_error = parse_positive_number(raw_actual_width, "actual width")
        if actual_error:
            return None, actual_error
        supplied_actual_width_mm = supplied_actual_width * width_scale_to_mm
        matching_width = next(
            (
                pair for pair in width_candidates
                if math.isclose(supplied_actual_width_mm, pair[1], rel_tol=0, abs_tol=0.000001)
            ),
            None,
        )
    elif len(width_candidates) == 1:
        matching_width = width_candidates[0]
    elif len(width_candidates) > 1:
        return None, f"actual width is required for nominal width {format_numeric_text(nominal_width_mm)} mm"
    else:
        matching_width = None
    if not matching_width:
        return None, f"invalid nominal/actual width for {blanket_name} at {format_numeric_text(thickness, 2)} mm"
    nominal_width, actual_width = matching_width

    raw_length = data.get("length")
    if parse_optional_text(raw_length) is None:
        raw_length = data.get("height")
    length, length_error = parse_positive_number(raw_length, "length")
    if length_error:
        return None, length_error
    length_unit = normalize_roll_length_unit(data.get("length_unit") or data.get("height_unit"))
    if not length_unit:
        return None, "length unit must be m, mm, or inch"

    roll_no = parse_optional_text(data.get("roll_no"))
    batch_no = parse_optional_text(data.get("batch_no"))

    if storage_type == M3Z_ROLL_STORAGE_TYPE:
        raw_rolls = data.get("number_of_rolls")
        if parse_optional_text(raw_rolls) is None:
            raw_rolls = data.get("rolls")
        if parse_optional_text(raw_rolls) is None and require_rolls:
            return None, "number of rolls is required"
        if parse_optional_text(raw_rolls) is None:
            number_of_rolls = None
        else:
            number_of_rolls, rolls_error = parse_positive_integer(raw_rolls, "number of rolls")
            if rolls_error:
                return None, rolls_error
        number_of_sheets = None
    else:
        raw_sheets = data.get("number_of_sheets")
        if parse_optional_text(raw_sheets) is None:
            raw_sheets = data.get("sheets")
        if parse_optional_text(raw_sheets) is None and require_rolls:
            return None, "number of sheets is required"
        if parse_optional_text(raw_sheets) is None:
            number_of_sheets = None
        else:
            number_of_sheets, sheets_error = parse_positive_integer(raw_sheets, "number of sheets")
            if sheets_error:
                return None, sheets_error
        number_of_rolls = None

    if number_of_rolls and number_of_rolls > 1 and (roll_no or batch_no):
        return None, "multiple rolls require separate rows when Roll No. or Batch No. is provided"

    raw_print_type = parse_optional_text(data.get("print_type"))
    print_options = rule.get("print_types", [])
    if print_options:
        print_type = normalize_blanket_print_type(raw_print_type)
        if print_type not in print_options:
            return None, f"print type for {blanket_name} must be P or W/O"
    else:
        if raw_print_type:
            return None, f"print type is not applicable to {blanket_name}"
        print_type = None

    supplied_unit = clean_text(data.get("unit"))
    if storage_type == M3Z_ROLL_STORAGE_TYPE:
        if supplied_unit and not normalize_roll_stock_unit(supplied_unit):
            return None, "unit must be m² for Rubber Blanket rolls"
    elif supplied_unit and not normalize_sheet_stock_unit(supplied_unit):
        return None, "unit must be sheets for Rubber Blanket cut pieces"

    length_meters = length * {"m": 1, "mm": 0.001, "inch": 0.0254}[length_unit]
    actual_width_meters = actual_width / 1000
    area_per_roll_sqm = actual_width_meters * length_meters
    total_quantity = area_per_roll_sqm * (number_of_rolls or 0) if storage_type == M3Z_ROLL_STORAGE_TYPE else (number_of_sheets or 0)

    return {
        "blanket_name": blanket_name,
        "brand": blanket_name,
        "type": print_type or "__none__",
        "nominal_width": format_numeric_text(nominal_width),
        "actual_width": format_numeric_text(actual_width),
        "width": format_numeric_text(nominal_width),
        "width_unit": "mm",
        "actual_width_meters": round(actual_width_meters, 9),
        "height": format_numeric_text(length),
        "length": format_numeric_text(length),
        "length_unit": length_unit,
        "length_meters": round(length_meters, 9),
        "thickness": f"{thickness:.2f}",
        "thickness_unit": "mm",
        "roll_no": roll_no,
        "batch_no": batch_no,
        "print_type": print_type,
        "number_of_rolls": number_of_rolls,
        "number_of_sheets": number_of_sheets,
        "storage_type": storage_type,
        "area_per_roll_sqm": round(area_per_roll_sqm, 6) if storage_type == M3Z_ROLL_STORAGE_TYPE else None,
        "area_per_sheet_sqm": round(area_per_roll_sqm, 6),
        "total_area_sqm": round(total_quantity, 6) if storage_type == M3Z_ROLL_STORAGE_TYPE else None,
        "total_quantity": round(total_quantity, 6),
        "unit": RUBBER_BLANKET_STOCK_UNIT if storage_type == M3Z_ROLL_STORAGE_TYPE else M3Z_SHEET_STOCK_UNIT,
    }, None


def parse_optional_text(value):
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        return clean_text(value)
    text = str(value).strip()
    return text or None


def parse_location_id(value):
    parsed = parse_optional_text(value)
    if parsed is None:
        return None, None
    location_id = parsed.upper()
    if not LOCATION_ID_PATTERN.fullmatch(location_id):
        return None, "location_id must use the format WH-A-RACK-L1-A"
    return location_id, None


def category_requires_brand(category):
    return category not in NO_BRAND_TYPE_CATEGORIES


def category_requires_type(category):
    return category not in NO_BRAND_TYPE_CATEGORIES


def normalize_optional_dimension(value):
    parsed = parse_optional_text(value)
    if parsed is None:
        return None
    compact = re.sub(r"\s+", "", parsed)
    return compact


def normalize_rule_type(value):
    cleaned = clean_text(value) or ""
    lowered = cleaned.lower()
    if lowered in {"packet", "pack", "pkt"}:
        return "pkt"
    if lowered in {"coil", "coils"}:
        return "coil"
    return cleaned


def parse_format_type(value):
    cleaned = clean_text(value)
    if not cleaned:
        return None, None, None, "type is required"

    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(ltr|l|kg|g|ml)", cleaned.strip(), flags=re.IGNORECASE)
    if not match:
        return None, None, None, "type must be a format like 1ltr, 5 ltr, 1kg"

    amount = float(match.group(1))
    unit_raw = match.group(2).lower()
    if unit_raw == "l":
        unit_raw = "ltr"
    normalized_type = f"{match.group(1)} {unit_raw}".replace("  ", " ")
    return normalized_type, amount, unit_raw, None


def category_requires_thickness(category):
    return category in THICKNESS_REQUIRED_CATEGORIES


def category_uses_dimensions(category):
    return category in DIMENSIONAL_CATEGORIES


def normalize_dimension(value):
    return normalize_optional_dimension(value)


def build_size_label(width, length):
    if width and length:
        return f"{width} x {length}"
    return width or length or None


def requires_batch_roll_no(category, unit):
    cleaned_unit = clean_text(unit)
    return category in BLANKET_BATCH_ROLL_CATEGORIES and cleaned_unit and cleaned_unit.lower() == "rolls"


def is_roll_unit(unit):
    cleaned_unit = clean_text(unit)
    return cleaned_unit is not None and cleaned_unit.lower() == "rolls"


def normalize_size(value):
    cleaned = parse_optional_text(value)
    if not cleaned:
        return None
    match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*", cleaned)
    if not match:
        return None
    return f"{match.group(1)} X {match.group(2)}"


def split_size(value):
    normalized = normalize_size(value)
    if not normalized:
        return None, None
    width, height = normalized.split(" X ", 1)
    return width, height


def size_from_data(data, width, height):
    explicit_size = normalize_size(data.get("size"))
    if explicit_size:
        return explicit_size
    if width and height:
        return normalize_size(f"{width} X {height}")
    return None


def normalize_creasing_thickness(value):
    number = parse_dimension_number(value)
    if number is None or not float(number).is_integer():
        return None
    return str(int(number))


def normalize_ctcp_thickness(value):
    number = parse_dimension_number(value)
    if number is None:
        return None
    if abs(number - 0.30) < 0.000001:
        return "0.30"
    if abs(number - 0.20) < 0.000001:
        return "0.20"
    return None


def find_chemical_product(value):
    cleaned = clean_text(value)
    if not cleaned:
        return None, None
    lowered = cleaned.casefold()
    for product_name, config in SPECIALIZED_CHEMICAL_PRODUCTS.items():
        if product_name.casefold() == lowered:
            return product_name, config
    return None, None


def parse_optional_integer(value, field_name):
    if parse_optional_text(value) is None and not isinstance(value, (int, float)):
        return None, None
    try:
        if bool(pd.isna(value)):
            return None, None
    except (TypeError, ValueError):
        pass
    return parse_integer(value, field_name)


def quantities_match(left, right):
    return abs(float(left) - float(right)) < 0.000001


def canonical_chemical_type(config, containers_per_box):
    amount = format(config["pack_size"], "g")
    base = f"{amount} {config['unit']}"
    if len(config["containers_per_box"]) > 1:
        return f"{base} / {containers_per_box} per box"
    return base


def infer_chemical_containers_per_box(item_type, config):
    cleaned = clean_text(item_type) or ""
    match = re.search(r"/\s*(\d+)\s+per\s+box$", cleaned, flags=re.IGNORECASE)
    if match:
        selected = int(match.group(1))
        if selected in config["containers_per_box"]:
            return selected
    return config["containers_per_box"][0]


def build_packaging(kind, **values):
    return {"kind": kind, **values}


def build_item_payload(data):
    category = normalize_category_name(data.get("category"))
    brand = clean_text(data.get("brand"))
    item_type = clean_text(data.get("type"))
    batch_roll_no = parse_optional_text(data.get("batch_roll_no"))
    width = normalize_dimension(data.get("width"))
    height = normalize_dimension(data.get("height"))
    thickness = parse_optional_text(data.get("thickness"))
    unit = clean_text(data.get("unit"))
    location_id, location_error = parse_location_id(data.get("location_id"))
    if location_error:
        return None, location_error
    if not category:
        return None, "category is required"
    if category not in ALLOWED_CATEGORIES:
        return None, "unsupported category"

    if category in NO_BRAND_TYPE_CATEGORIES:
        brand = "__none__"
        item_type = "__none__"

    format_size = None
    format_unit = None
    packaging = None
    roll_paper_fields = None
    rubber_blanket_fields = None

    if category == "Creasing Matrix":
        thickness = normalize_creasing_thickness(thickness)
        if thickness not in CREASING_MATRIX_SIZES:
            return None, "Creasing Matrix thickness must be 9 mm, 11 mm, or 13 mm"
        selected_size = size_from_data(data, width, height)
        if not selected_size:
            return None, "Creasing Matrix size is required in the format 0.4 X 1.5"
        if selected_size not in CREASING_MATRIX_SIZES[thickness]:
            return None, f"size {selected_size} is not valid for {thickness} mm Creasing Matrix"
        width, height = split_size(selected_size)
        if not unit or unit.lower() not in {"pkt", "packet", "packets"}:
            return None, "unit must be pkt for Creasing Matrix"
        unit = "pkt"
        boxes, boxes_error = parse_optional_integer(data.get("boxes"), "boxes")
        loose_units, loose_error = parse_optional_integer(data.get("loose_units"), "loose units")
        if boxes_error or loose_error:
            return None, boxes_error or loose_error
        if boxes is not None or loose_units is not None:
            boxes = boxes or 0
            loose_units = loose_units or 0
            if loose_units >= 10:
                return None, "loose packets must be less than 10"
            quantity = boxes * 10 + loose_units
            if parse_optional_text(data.get("quantity")) is not None:
                supplied_quantity, quantity_error = parse_integer(data.get("quantity"), "quantity")
                if quantity_error:
                    return None, quantity_error
                if supplied_quantity != quantity:
                    return None, "quantity does not match boxes and loose packets"
        else:
            quantity, quantity_error = parse_integer(data.get("quantity"), "quantity")
            if quantity_error:
                return None, quantity_error
        packaging = build_packaging("creasing_matrix", units_per_box=10, container_type="packet")
        batch_roll_no = None

    elif category == "CTCP Plates":
        thickness = normalize_ctcp_thickness(thickness)
        if thickness not in CTCP_PLATE_SIZES:
            return None, "CTCP Plates thickness must be 0.30 or 0.20"
        selected_size = size_from_data(data, width, height)
        if not selected_size:
            return None, "CTCP Plates size is required in the format 650 X 550"
        if selected_size not in CTCP_PLATE_SIZES[thickness]:
            return None, f"size {selected_size} is not valid for {thickness} CTCP Plates"
        width, height = split_size(selected_size)
        if not unit or unit.lower() not in {"box", "boxes"}:
            return None, "unit must be box for CTCP Plates"
        unit = "box"
        boxes, boxes_error = parse_optional_integer(data.get("boxes"), "boxes")
        if boxes_error:
            return None, boxes_error
        if boxes is not None:
            quantity = boxes
            if parse_optional_text(data.get("quantity")) is not None:
                supplied_quantity, quantity_error = parse_integer(data.get("quantity"), "quantity")
                if quantity_error:
                    return None, quantity_error
                if supplied_quantity != quantity:
                    return None, "quantity does not match CTCP box count"
        else:
            quantity, quantity_error = parse_integer(data.get("quantity"), "quantity")
            if quantity_error:
                return None, quantity_error
        if parse_optional_text(data.get("total_sheets")) is not None:
            supplied_sheets, sheets_error = parse_integer(data.get("total_sheets"), "total sheets")
            if sheets_error:
                return None, sheets_error
            if supplied_sheets != quantity * 50:
                return None, "total sheets must equal CTCP boxes multiplied by 50"
        packaging = build_packaging("ctcp_plates", sheets_per_box=50, container_type="box")
        batch_roll_no = None

    elif category == RUBBER_BLANKET_CATEGORY:
        rubber_blanket_fields, rubber_blanket_error = build_rubber_blanket_fields(data)
        if rubber_blanket_error:
            return None, rubber_blanket_error
        brand = rubber_blanket_fields["brand"]
        item_type = rubber_blanket_fields["type"]
        batch_roll_no = None
        width = rubber_blanket_fields["width"]
        height = rubber_blanket_fields["height"]
        thickness = rubber_blanket_fields["thickness"]
        quantity = rubber_blanket_fields["total_quantity"]
        unit = rubber_blanket_fields["unit"]
        rubber_stock_unit = rubber_blanket_fields["unit"]
        packaging = build_packaging(
            "rubber_blankets",
            stock_unit=RUBBER_BLANKET_STOCK_UNIT,
            movement_units=["rolls", "m²"],
        )
        packaging["stock_unit"] = rubber_stock_unit
        packaging["movement_units"] = (
            ["rolls", RUBBER_BLANKET_STOCK_UNIT]
            if rubber_blanket_fields["storage_type"] == M3Z_ROLL_STORAGE_TYPE
            else ["sheets"]
        )
        packaging["storage_type"] = rubber_blanket_fields["storage_type"]

    elif category == ROLL_PAPER_CATEGORY:
        roll_paper_fields, roll_paper_error = build_roll_paper_fields(data)
        if roll_paper_error:
            return None, roll_paper_error
        width = roll_paper_fields["width"]
        height = roll_paper_fields["height"]
        thickness = roll_paper_fields["thickness"]
        quantity = roll_paper_fields["total_quantity"]
        unit = roll_paper_fields["unit"]
        batch_roll_no = None
        packaging = build_packaging(
            "calibrated_underpacking_paper",
            stock_unit=roll_paper_fields["unit"],
            movement_units=["rolls"] if roll_paper_fields["storage_type"] == M3Z_ROLL_STORAGE_TYPE else ["sheets"],
        )

    else:
        product_name, product_config = find_chemical_product(data.get("product") or brand)
        if category in CHEMICAL_CATEGORIES and product_config:
            brand = product_name
        requires_brand = category_requires_brand(category)
        requires_type = category_requires_type(category)
        if requires_brand and not brand:
            return None, "brand is required"
        if requires_type and not item_type and not product_config:
            return None, "type is required"
        if not unit:
            return None, "unit is required"
        if requires_batch_roll_no(category, unit) and not batch_roll_no:
            return None, "batch / roll no. is required for blanket rolls"
        if not requires_batch_roll_no(category, unit):
            batch_roll_no = None

        if category in RULE_UNIT_LINKED_CATEGORIES:
            normalized_type = normalize_rule_type(item_type)
            if normalized_type not in {"coil", "pkt"}:
                return None, "type must be coil or pkt for this category"
            item_type = normalized_type
            if unit.lower() != item_type:
                return None, "unit must match type for this category"

        if category in CHEMICAL_CATEGORIES and product_config:
            if category != product_config["category"]:
                return None, f"{product_name} belongs to {product_config['category']}"
            brand = product_name
            selected_per_box, per_box_error = parse_optional_integer(
                data.get("containers_per_box"), "containers per box"
            )
            if per_box_error:
                return None, per_box_error
            if selected_per_box is None:
                selected_per_box = infer_chemical_containers_per_box(item_type, product_config)
            if selected_per_box not in product_config["containers_per_box"]:
                allowed = ", ".join(str(value) for value in product_config["containers_per_box"])
                return None, f"containers per box must be one of: {allowed}"

            if parse_optional_text(data.get("pack_size")) is not None:
                supplied_pack_size, pack_size_error = parse_number(data.get("pack_size"), "pack size")
                if pack_size_error:
                    return None, pack_size_error
                if not quantities_match(supplied_pack_size, product_config["pack_size"]):
                    return None, f"pack size for {product_name} must be {product_config['pack_size']}"
            supplied_container_type = clean_text(data.get("container_type"))
            if supplied_container_type and supplied_container_type.lower() != product_config["container_type"]:
                return None, f"container type for {product_name} must be {product_config['container_type']}"

            expected_type = canonical_chemical_type(product_config, selected_per_box)
            if item_type:
                base_type = item_type.split("/", 1)[0].strip()
                _, supplied_pack_size, supplied_pack_unit, format_error = parse_format_type(base_type)
                if format_error or not quantities_match(supplied_pack_size, product_config["pack_size"]):
                    return None, f"pack size for {product_name} must be {product_config['pack_size']} {product_config['unit']}"
                if supplied_pack_unit != product_config["unit"]:
                    return None, f"pack unit for {product_name} must be {product_config['unit']}"
            item_type = expected_type
            unit = unit.lower()
            if unit == "l":
                unit = "ltr"
            if unit != product_config["unit"]:
                return None, f"unit for {product_name} must be {product_config['unit']}"

            boxes, boxes_error = parse_optional_integer(data.get("boxes"), "boxes")
            loose_units, loose_error = parse_optional_integer(data.get("loose_units"), "loose containers")
            containers, containers_error = parse_optional_integer(data.get("containers"), "containers")
            if boxes_error or loose_error or containers_error:
                return None, boxes_error or loose_error or containers_error

            if boxes is not None or loose_units is not None:
                boxes = boxes or 0
                loose_units = loose_units or 0
                if loose_units >= selected_per_box:
                    return None, f"loose containers must be less than {selected_per_box}"
                calculated_containers = boxes * selected_per_box + loose_units
                if containers is not None and containers != calculated_containers:
                    return None, "container count does not match boxes and loose containers"
                containers = calculated_containers

            if containers is not None:
                quantity = round(containers * product_config["pack_size"], 4)
                if parse_optional_text(data.get("quantity")) is not None:
                    supplied_quantity, quantity_error = parse_number(data.get("quantity"), "quantity")
                    if quantity_error:
                        return None, quantity_error
                    if not quantities_match(supplied_quantity, quantity):
                        return None, "quantity does not match the chemical container calculation"
            else:
                quantity, quantity_error = parse_number(data.get("quantity"), "quantity")
                if quantity_error:
                    return None, quantity_error
                containers = quantity / product_config["pack_size"]
                if not quantities_match(containers, round(containers)):
                    return None, "chemical quantity must equal a whole number of containers"

            format_size = product_config["pack_size"]
            format_unit = product_config["unit"]
            packaging = build_packaging(
                "chemical",
                pack_size=product_config["pack_size"],
                pack_unit=product_config["unit"],
                container_type=product_config["container_type"],
                containers_per_box=selected_per_box,
                display_format=product_config["display_format"],
            )
        else:
            if category in CHEMICAL_CATEGORIES:
                normalized_type, format_size, format_unit, format_error = parse_format_type(item_type)
                if format_error:
                    return None, format_error
                item_type = normalized_type
                if unit.lower() != format_unit:
                    return None, "unit must match the type format unit (e.g., ltr or kg)"

            if is_roll_unit(unit):
                quantity = 0
                quantity_error = None
            elif category in CHEMICAL_CATEGORIES:
                quantity, quantity_error = parse_number(data.get("quantity"), "quantity")
            else:
                quantity, quantity_error = parse_integer(data.get("quantity"), "quantity")
            if quantity_error:
                return None, quantity_error

        if category_uses_dimensions(category) and not all([width, height]):
            return None, "width and length are required for this category"
        if category_requires_thickness(category) and not thickness:
            return None, "thickness is required for this category"
        if is_roll_unit(unit):
            roll_area = calculate_roll_area_sqm(width, height)
            if roll_area is None:
                return None, "width and length must be numeric for roll sq.m calculation"
            quantity = roll_area

    now = now_ist()
    payload = {
        "category": category,
        "brand": brand,
        "type": item_type,
        "batch_roll_no": batch_roll_no,
        "width": width,
        "height": height,
        "size": build_size_label(width, height),
        "thickness": thickness,
        "quantity": quantity,
        "unit": unit,
        "created_at": now,
        "updated_at": now,
    }
    if location_id is not None:
        payload["location_id"] = location_id
    if format_size is not None:
        payload["format_size"] = format_size
        payload["format_unit"] = format_unit
    if packaging is not None:
        payload["packaging"] = packaging
    if roll_paper_fields is not None:
        payload.update(
            {
                "width_unit": roll_paper_fields["width_unit"],
                "storage_type": roll_paper_fields["storage_type"],
                "length": roll_paper_fields["length"],
                "length_unit": roll_paper_fields["length_unit"],
                "thickness_unit": roll_paper_fields["thickness_unit"],
                "thickness_micron": roll_paper_fields["thickness_micron"],
                "number_of_rolls": roll_paper_fields["number_of_rolls"],
                "number_of_sheets": roll_paper_fields["number_of_sheets"],
                "width_meters": roll_paper_fields["width_meters"],
                "length_meters": roll_paper_fields["length_meters"],
                "area_per_roll_sqm": roll_paper_fields["area_per_roll_sqm"],
                "area_per_sheet_sqm": roll_paper_fields["area_per_sheet_sqm"],
            }
        )
    if rubber_blanket_fields is not None:
        payload.update(
            {
                "blanket_name": rubber_blanket_fields["blanket_name"],
                "nominal_width": rubber_blanket_fields["nominal_width"],
                "actual_width": rubber_blanket_fields["actual_width"],
                "width_unit": rubber_blanket_fields["width_unit"],
                "actual_width_meters": rubber_blanket_fields["actual_width_meters"],
                "length": rubber_blanket_fields["length"],
                "length_unit": rubber_blanket_fields["length_unit"],
                "length_meters": rubber_blanket_fields["length_meters"],
                "thickness_unit": rubber_blanket_fields["thickness_unit"],
                "roll_no": rubber_blanket_fields["roll_no"],
                "batch_no": rubber_blanket_fields["batch_no"],
                "print_type": rubber_blanket_fields["print_type"],
                "storage_type": rubber_blanket_fields["storage_type"],
                "number_of_rolls": rubber_blanket_fields["number_of_rolls"],
                "number_of_sheets": rubber_blanket_fields["number_of_sheets"],
                "area_per_roll_sqm": rubber_blanket_fields["area_per_roll_sqm"],
                "area_per_sheet_sqm": rubber_blanket_fields["area_per_sheet_sqm"],
            }
        )
    return payload, None


def build_lookup(data):
    category = normalize_category_name(data.get("category"))
    brand = clean_text(data.get("brand"))
    item_type = clean_text(data.get("type"))
    batch_roll_no = parse_optional_text(data.get("batch_roll_no"))
    width = normalize_dimension(data.get("width"))
    height = normalize_dimension(data.get("height"))
    thickness = parse_optional_text(data.get("thickness"))
    unit = clean_text(data.get("unit"))

    if not category:
        return None, "category is required"
    if category not in ALLOWED_CATEGORIES:
        return None, "unsupported category"

    if category in NO_BRAND_TYPE_CATEGORIES:
        brand = "__none__"
        item_type = "__none__"

    if category == "Creasing Matrix":
        thickness = normalize_creasing_thickness(thickness)
        selected_size = size_from_data(data, width, height)
        if thickness not in CREASING_MATRIX_SIZES:
            return None, "Creasing Matrix thickness must be 9 mm, 11 mm, or 13 mm"
        if not selected_size or selected_size not in CREASING_MATRIX_SIZES[thickness]:
            return None, "invalid Creasing Matrix thickness and size combination"
        width, height = split_size(selected_size)

    if category == "CTCP Plates":
        thickness = normalize_ctcp_thickness(thickness)
        selected_size = size_from_data(data, width, height)
        if thickness not in CTCP_PLATE_SIZES:
            return None, "CTCP Plates thickness must be 0.30 or 0.20"
        if not selected_size or selected_size not in CTCP_PLATE_SIZES[thickness]:
            return None, "invalid CTCP Plates thickness and size combination"
        width, height = split_size(selected_size)

    rubber_blanket_fields = None
    if category == RUBBER_BLANKET_CATEGORY:
        explicit_blanket_name = parse_optional_text(data.get("blanket_name"))
        canonical_brand, _ = find_rubber_blanket_rule(brand)
        if explicit_blanket_name or canonical_brand:
            rubber_blanket_fields, rubber_blanket_error = build_rubber_blanket_fields(data, require_rolls=False)
            if rubber_blanket_error:
                return None, rubber_blanket_error
            brand = rubber_blanket_fields["brand"]
            item_type = rubber_blanket_fields["type"]
            batch_roll_no = None
            width = rubber_blanket_fields["width"]
            height = rubber_blanket_fields["height"]
            thickness = rubber_blanket_fields["thickness"]
            unit = rubber_blanket_fields["unit"]

    roll_paper_fields = None
    if category == ROLL_PAPER_CATEGORY:
        roll_paper_fields, roll_paper_error = build_roll_paper_fields(data, require_rolls=False)
        if roll_paper_error:
            return None, roll_paper_error
        width = roll_paper_fields["width"]
        height = roll_paper_fields["height"]
        thickness = roll_paper_fields["thickness"]
        unit = roll_paper_fields["unit"]
        brand = "__none__"
        item_type = "__none__"

    if category in CHEMICAL_CATEGORIES:
        product_name, product_config = find_chemical_product(data.get("product") or brand)
        if product_config:
            if category != product_config["category"]:
                return None, f"{product_name} belongs to {product_config['category']}"
            brand = product_name
            containers_per_box = infer_chemical_containers_per_box(item_type, product_config)
            item_type = canonical_chemical_type(product_config, containers_per_box)

    if category_requires_brand(category) and not brand:
        return None, "brand is required"
    if category_requires_type(category) and not item_type:
        return None, "type is required"
    if not requires_batch_roll_no(category, unit):
        batch_roll_no = None

    if category_uses_dimensions(category) and not all([width, height]):
        return None, "width and length are required for this category"

    if category_requires_thickness(category) and not thickness:
        return None, "thickness is required for this category"

    lookup = {
        "category": category,
        "brand": brand,
        "type": item_type,
        "batch_roll_no": batch_roll_no,
        "width": width,
        "height": height,
        "thickness": thickness,
    }
    if roll_paper_fields is not None:
        lookup.update(
            {
                "width_unit": roll_paper_fields["width_unit"],
                "storage_type": roll_paper_fields["storage_type"],
                "length": roll_paper_fields["length"],
                "length_unit": roll_paper_fields["length_unit"],
                "thickness_unit": roll_paper_fields["thickness_unit"],
                "thickness_micron": roll_paper_fields["thickness_micron"],
                "width_meters": roll_paper_fields["width_meters"],
                "length_meters": roll_paper_fields["length_meters"],
                "number_of_rolls": roll_paper_fields["number_of_rolls"],
                "number_of_sheets": roll_paper_fields["number_of_sheets"],
            }
        )
    if rubber_blanket_fields is not None:
        lookup.update(
            {
                "blanket_name": rubber_blanket_fields["blanket_name"],
                "nominal_width": rubber_blanket_fields["nominal_width"],
                "actual_width": rubber_blanket_fields["actual_width"],
                "actual_width_meters": rubber_blanket_fields["actual_width_meters"],
                "width_unit": rubber_blanket_fields["width_unit"],
                "length": rubber_blanket_fields["length"],
                "length_unit": rubber_blanket_fields["length_unit"],
                "length_meters": rubber_blanket_fields["length_meters"],
                "thickness_unit": rubber_blanket_fields["thickness_unit"],
                "roll_no": rubber_blanket_fields["roll_no"],
                "batch_no": rubber_blanket_fields["batch_no"],
                "print_type": rubber_blanket_fields["print_type"],
                "storage_type": (
                    {"$in": [M3Z_ROLL_STORAGE_TYPE, None]}
                    if rubber_blanket_fields["storage_type"] == M3Z_ROLL_STORAGE_TYPE
                    else rubber_blanket_fields["storage_type"]
                ),
            }
        )
    return lookup, None


def build_item_key(lookup):
    parts = [
        lookup["category"],
        lookup["brand"],
        lookup["type"],
        lookup.get("batch_roll_no") or "-",
    ]
    if lookup.get("category") == RUBBER_BLANKET_CATEGORY:
        parts.extend(
            [
                lookup.get("blanket_name") or lookup.get("brand") or "-",
                lookup.get("storage_type") or M3Z_ROLL_STORAGE_TYPE,
                format_numeric_text(lookup.get("actual_width_meters", 0), 9),
                format_numeric_text(lookup.get("length_meters", 0), 9),
                lookup.get("thickness") or "-",
                lookup.get("roll_no") or "-",
                lookup.get("batch_no") or "-",
                lookup.get("print_type") or "-",
            ]
        )
    elif lookup.get("category") == ROLL_PAPER_CATEGORY:
        parts.extend(
            [
                lookup.get("storage_type") or M3Z_ROLL_STORAGE_TYPE,
                format_numeric_text(lookup.get("width_meters", 0), 9),
                format_numeric_text(lookup.get("length_meters", 0), 9),
                format_numeric_text(lookup.get("thickness_micron", 0), 6),
            ]
        )
    else:
        parts.extend(
            [
                lookup.get("width") or "-",
                lookup.get("height") or "-",
                lookup.get("thickness") or "-",
            ]
        )
    return "|".join(parts)


def infer_item_packaging(item):
    packaging = item.get("packaging")
    if isinstance(packaging, dict) and packaging.get("kind"):
        return dict(packaging)
    category = item.get("category")
    if category == "Creasing Matrix":
        return build_packaging("creasing_matrix", units_per_box=10, container_type="packet")
    if category == "CTCP Plates":
        return build_packaging("ctcp_plates", sheets_per_box=50, container_type="box")
    if category == RUBBER_BLANKET_CATEGORY and item.get("blanket_name"):
        storage_type = item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE
        packaging = build_packaging(
            "rubber_blankets",
            stock_unit=RUBBER_BLANKET_STOCK_UNIT,
            movement_units=["rolls", "m²"],
        )
        packaging["stock_unit"] = RUBBER_BLANKET_STOCK_UNIT if storage_type == M3Z_ROLL_STORAGE_TYPE else M3Z_SHEET_STOCK_UNIT
        packaging["movement_units"] = ["rolls", RUBBER_BLANKET_STOCK_UNIT] if storage_type == M3Z_ROLL_STORAGE_TYPE else ["sheets"]
        packaging["storage_type"] = storage_type
        return packaging
    if category == ROLL_PAPER_CATEGORY:
        return build_packaging(
            "calibrated_underpacking_paper",
            stock_unit=(M3Z_SHEET_STOCK_UNIT if item.get("storage_type") == M3Z_CUT_PIECE_STORAGE_TYPE else ROLL_PAPER_STOCK_UNIT),
            movement_units=["sheets"] if item.get("storage_type") == M3Z_CUT_PIECE_STORAGE_TYPE else ["rolls"],
        )
    if category in CHEMICAL_CATEGORIES:
        _, product_config = find_chemical_product(item.get("brand"))
        if product_config and category == product_config["category"]:
            return build_packaging(
                "chemical",
                pack_size=product_config["pack_size"],
                pack_unit=product_config["unit"],
                container_type=product_config["container_type"],
                containers_per_box=infer_chemical_containers_per_box(item.get("type"), product_config),
                display_format=product_config["display_format"],
            )
    return None


def build_stock_breakdown(item):
    packaging = infer_item_packaging(item)
    if not packaging:
        return None
    quantity = float(item.get("quantity", 0) or 0)
    kind = packaging["kind"]
    if kind == "creasing_matrix":
        units_per_box = packaging["units_per_box"]
        packet_quantity = int(round(quantity))
        return {
            "packets": packet_quantity,
            "boxes": packet_quantity // units_per_box,
            "loose_units": packet_quantity % units_per_box,
        }
    if kind == "ctcp_plates":
        box_quantity = int(round(quantity))
        return {
            "boxes": box_quantity,
            "total_sheets": box_quantity * packaging["sheets_per_box"],
        }
    if kind == "chemical":
        containers = quantity / packaging["pack_size"]
        rounded_containers = int(round(containers)) if quantities_match(containers, round(containers)) else containers
        containers_per_box = packaging["containers_per_box"]
        if isinstance(rounded_containers, int):
            boxes = rounded_containers // containers_per_box
            loose_units = rounded_containers % containers_per_box
        else:
            boxes = None
            loose_units = None
        return {
            "containers": rounded_containers,
            "boxes": boxes,
            "loose_units": loose_units,
            "normalized_quantity": quantity,
            "normalized_unit": item.get("unit"),
        }
    if kind == "rubber_blankets":
        area_per_roll = item.get("area_per_roll_sqm")
        if area_per_roll is None:
            try:
                actual_width_meters = float(item.get("actual_width")) / 1000
                length_unit = normalize_roll_length_unit(item.get("length_unit"))
                length_meters = float(item.get("length", item.get("height"))) * {
                    "m": 1,
                    "mm": 0.001,
                    "inch": 0.0254,
                }[length_unit]
            except (KeyError, TypeError, ValueError):
                return None
            area_per_roll = actual_width_meters * length_meters
        if (item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE) == M3Z_CUT_PIECE_STORAGE_TYPE:
            return {
                "sheets": item.get("number_of_sheets"),
                "area_per_sheet_sqm": round(float(item.get("area_per_sheet_sqm") or area_per_roll), 6),
                "total_sheets": quantity,
                "normalized_quantity": quantity,
                "normalized_unit": item.get("unit") or M3Z_SHEET_STOCK_UNIT,
            }
        return {
            "rolls": item.get("number_of_rolls"),
            "area_per_roll_sqm": round(float(area_per_roll), 6),
            "total_area_sqm": round(quantity, 6),
            "normalized_quantity": quantity,
            "normalized_unit": item.get("unit") or RUBBER_BLANKET_STOCK_UNIT,
        }
    if kind == "calibrated_underpacking_paper":
        storage_type = item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE
        if storage_type == M3Z_CUT_PIECE_STORAGE_TYPE:
            area_per_sheet = item.get("area_per_sheet_sqm")
            if area_per_sheet is None:
                try:
                    width_meters = float(item.get("width_meters"))
                    length_meters = float(item.get("length_meters"))
                except (TypeError, ValueError):
                    return None
                area_per_sheet = width_meters * length_meters
            sheets = item.get("number_of_sheets")
            return {
                "sheets": int(sheets) if sheets is not None else None,
                "area_per_sheet_sqm": round(float(area_per_sheet), 6),
                "total_sheets": quantity,
                "normalized_quantity": quantity,
                "normalized_unit": item.get("unit") or M3Z_SHEET_STOCK_UNIT,
            }
        area_per_roll = item.get("area_per_roll_sqm")
        if area_per_roll is None:
            width_meters = item.get("width_meters")
            length_meters = item.get("length_meters")
            if width_meters is None or length_meters is None:
                width_unit = normalize_roll_width_unit(item.get("width_unit"))
                length_unit = normalize_roll_length_unit(item.get("length_unit"))
                try:
                    width_meters = float(item.get("width")) * {"mm": 0.001, "m": 1, "inch": 0.0254}[width_unit]
                    length_meters = float(item.get("length", item.get("height"))) * {"m": 1, "mm": 0.001, "inch": 0.0254}[length_unit]
                except (KeyError, TypeError, ValueError):
                    return None
            area_per_roll = width_meters * length_meters
        rolls = item.get("number_of_rolls")
        return {
            "rolls": int(rolls) if rolls is not None else None,
            "area_per_roll_sqm": round(float(area_per_roll), 6),
            "total_area_sqm": round(quantity, 6),
            "normalized_quantity": quantity,
            "normalized_unit": item.get("unit") or ROLL_PAPER_STOCK_UNIT,
        }
    return None


def serialize_item(item):
    packaging = infer_item_packaging(item)
    serialized = {
        "id": str(item["_id"]),
        "category": item["category"],
        "brand": item["brand"],
        "type": item["type"],
        "batch_roll_no": item.get("batch_roll_no"),
        "width": item.get("width"),
        "height": item.get("height"),
        "size": item.get("size"),
        "thickness": item.get("thickness"),
        "quantity": item["quantity"],
        "unit": item["unit"],
        "location_id": item.get("location_id"),
        "packaging": packaging,
        "stock_breakdown": build_stock_breakdown(item),
        "created_at": serialize_datetime_ist(item.get("created_at")),
        "updated_at": serialize_datetime_ist(item.get("updated_at")),
    }
    if item.get("category") == ROLL_PAPER_CATEGORY:
        serialized.update(
            {
                "storage_type": item.get("storage_type", M3Z_ROLL_STORAGE_TYPE),
                "width_unit": item.get("width_unit"),
                "length": item.get("length", item.get("height")),
                "length_unit": item.get("length_unit"),
                "thickness_unit": item.get("thickness_unit"),
                "thickness_micron": item.get("thickness_micron"),
                "number_of_rolls": item.get("number_of_rolls"),
                "number_of_sheets": item.get("number_of_sheets"),
                "area_per_roll_sqm": item.get("area_per_roll_sqm"),
                "area_per_sheet_sqm": item.get("area_per_sheet_sqm"),
            }
        )
    if item.get("category") == RUBBER_BLANKET_CATEGORY:
        canonical_blanket_name, _ = find_rubber_blanket_rule(item.get("blanket_name") or item.get("brand"))
        if canonical_blanket_name:
            serialized["brand"] = canonical_blanket_name
        serialized.update(
            {
                "blanket_name": canonical_blanket_name or item.get("blanket_name", item.get("brand")),
                "storage_type": item.get("storage_type", M3Z_ROLL_STORAGE_TYPE),
                "nominal_width": item.get("nominal_width", item.get("width")),
                "actual_width": item.get("actual_width", item.get("width")),
                "width_unit": item.get("width_unit", "mm"),
                "length": item.get("length", item.get("height")),
                "length_unit": item.get("length_unit", "m"),
                "thickness_unit": item.get("thickness_unit", "mm"),
                "roll_no": item.get("roll_no"),
                "batch_no": item.get("batch_no"),
                "print_type": item.get("print_type"),
                "number_of_rolls": item.get("number_of_rolls"),
                "number_of_sheets": item.get("number_of_sheets"),
                "area_per_roll_sqm": item.get("area_per_roll_sqm"),
                "area_per_sheet_sqm": item.get("area_per_sheet_sqm"),
            }
        )
    return serialized


def serialize_log(log):
    return {
        "id": str(log["_id"]),
        "item_key": log["item_key"],
        "action": log["action"],
        "category": log["category"],
        "brand": log["brand"],
        "type": log["type"],
        "batch_roll_no": log.get("batch_roll_no"),
        "blanket_name": log.get("blanket_name"),
        "roll_no": log.get("roll_no"),
        "batch_no": log.get("batch_no"),
        "print_type": log.get("print_type"),
        "size": log["size"],
        "width": log.get("width"),
        "height": log.get("height"),
        "thickness": log.get("thickness"),
        "thickness_unit": log.get("thickness_unit"),
        "thickness_micron": log.get("thickness_micron"),
        "storage_type": log.get("storage_type"),
        "number_of_rolls": log.get("number_of_rolls"),
        "number_of_sheets": log.get("number_of_sheets"),
        "area_per_roll_sqm": log.get("area_per_roll_sqm"),
        "area_per_sheet_sqm": log.get("area_per_sheet_sqm"),
        "quantity_before": log["quantity_before"],
        "quantity_after": log["quantity_after"],
        "quantity_change": log["quantity_change"],
        "unit": log["unit"],
        "packaging": log.get("packaging"),
        "source": log["source"],
        "reason": log.get("reason"),
        "details": log.get("details") or {},
        "changed_at": serialize_datetime_ist(log.get("changed_at")),
    }


def log_stock_change(item, action, quantity_before, quantity_after, source, reason=None, details=None):
    stock_logs_collection = get_stock_logs_collection()
    stock_logs_collection.insert_one(
        {
            "item_key": build_item_key(item),
            "action": action,
            "category": item["category"],
            "brand": item["brand"],
            "type": item["type"],
            "batch_roll_no": item.get("batch_roll_no"),
            "blanket_name": item.get("blanket_name"),
            "roll_no": item.get("roll_no"),
            "batch_no": item.get("batch_no"),
            "print_type": item.get("print_type"),
            "size": item.get("size"),
            "width": item.get("width"),
            "height": item.get("height"),
            "thickness": item.get("thickness"),
            "thickness_unit": item.get("thickness_unit"),
            "thickness_micron": item.get("thickness_micron"),
            "storage_type": item.get("storage_type"),
            "number_of_rolls": item.get("number_of_rolls"),
            "number_of_sheets": item.get("number_of_sheets"),
            "area_per_roll_sqm": item.get("area_per_roll_sqm"),
            "area_per_sheet_sqm": item.get("area_per_sheet_sqm"),
            "quantity_before": quantity_before,
            "quantity_after": quantity_after,
            "quantity_change": quantity_after - quantity_before,
            "unit": item["unit"],
            "packaging": infer_item_packaging(item),
            "source": source,
            "reason": parse_optional_text(reason),
            "details": details or {},
            "changed_at": now_ist(),
        }
    )


def create_inventory_query(args):
    query = {}

    category = normalize_category_name(args.get("category"))
    brand = clean_text(args.get("brand"))
    item_type = clean_text(args.get("type"))
    search = clean_text(args.get("search"))
    low_stock = str(args.get("low_stock", "")).lower() in {"1", "true", "yes"}
    thickness = parse_optional_text(args.get("thickness"))

    if category:
        query["category"] = {"$regex": f"^{re.escape(category)}$", "$options": "i"}
    if brand:
        query["brand"] = {"$regex": f"^{re.escape(brand)}$", "$options": "i"}
    if item_type:
        query["type"] = {"$regex": f"^{re.escape(item_type)}$", "$options": "i"}
    if thickness:
        query["thickness"] = {"$regex": f"^{re.escape(thickness)}$", "$options": "i"}

    if search:
        regex = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"category": regex},
            {"brand": regex},
            {"type": regex},
            {"batch_roll_no": regex},
            {"blanket_name": regex},
            {"roll_no": regex},
            {"batch_no": regex},
        ]

    if low_stock:
        threshold_value, threshold_error = parse_integer(
            args.get("low_stock_threshold", DEFAULT_LOW_STOCK_THRESHOLD),
            "low_stock_threshold",
        )
        if threshold_error:
            return None, threshold_error
        query["quantity"] = {"$lte": threshold_value}

    return query, None


def process_excel_row(row):
    nominal_width = row.get("Nominal Width")
    if parse_optional_text(nominal_width) is None:
        nominal_width = row.get("Width")
    storage_type = row.get("Storage Type")
    thickness_micron = row.get("Thickness (Micron)")
    number_of_rolls = row.get("Number of Rolls")
    if parse_optional_text(number_of_rolls) is None:
        number_of_rolls = row.get("No. of Rolls")
    if parse_optional_text(number_of_rolls) is None:
        number_of_rolls = row.get("Rolls")
    number_of_sheets = row.get("Number of Sheets")
    if parse_optional_text(number_of_sheets) is None:
        number_of_sheets = row.get("No. of Sheets")
    if parse_optional_text(number_of_sheets) is None:
        number_of_sheets = row.get("Sheets")
    roll_no = row.get("Roll No")
    normalized_category = normalize_category_name(row.get("Category"))
    if normalized_category in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} and "Storage Type" in row and parse_optional_text(storage_type) is None:
        storage_type = "__missing__"
    if parse_optional_text(roll_no) is None and normalized_category == RUBBER_BLANKET_CATEGORY:
        roll_no = row.get("Batch/Roll No")
    item, error = build_item_payload(
        {
            "category": row.get("Category"),
            "brand": row.get("Brand"),
            "type": row.get("Type"),
            "blanket_name": row.get("Blanket Name"),
            "batch_roll_no": row.get("Batch/Roll No"),
            "width": nominal_width,
            "nominal_width": nominal_width,
            "actual_width": row.get("Actual Width"),
            "width_unit": row.get("Width Unit"),
            "height": row.get("Length", row.get("Height")),
            "length": row.get("Length"),
            "length_unit": row.get("Length Unit"),
            "size": row.get("Size"),
            "thickness": row.get("Thickness"),
            "thickness_unit": row.get("Thickness Unit"),
            "thickness_micron": thickness_micron,
            "storage_type": storage_type,
            "number_of_rolls": number_of_rolls,
            "rolls": number_of_rolls,
            "number_of_sheets": number_of_sheets,
            "sheets": number_of_sheets,
            "roll_no": roll_no,
            "batch_no": row.get("Batch No"),
            "print_type": row.get("Print Type"),
            "area_per_roll_sqm": row.get("Area per Roll"),
            "quantity": row.get("Quantity"),
            "unit": row.get("Unit"),
            "product": row.get("Product"),
            "containers_per_box": row.get("Containers/Box"),
            "boxes": row.get("Boxes"),
            "loose_units": row.get("Loose Units"),
            "containers": row.get("Containers"),
            "pack_size": row.get("Pack Size"),
            "container_type": row.get("Container Type"),
            "total_sheets": row.get("Total Sheets"),
        }
    )
    return item, error


def get_request_reason(data, default_reason=None):
    reason = parse_optional_text(data.get("reason"))
    return reason or default_reason


def get_excel_reason():
    return parse_optional_text(request.form.get("reason")) or "Excel upload"


def get_item_identity_query(item):
    query = {
        "category": item["category"],
        "brand": item["brand"],
        "type": item["type"],
        "batch_roll_no": item.get("batch_roll_no"),
    }
    if item.get("category") == RUBBER_BLANKET_CATEGORY:
        storage_type = item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE
        query.update(
            {
                "blanket_name": item.get("blanket_name"),
                "storage_type": (
                    {"$in": [M3Z_ROLL_STORAGE_TYPE, None]}
                    if storage_type == M3Z_ROLL_STORAGE_TYPE
                    else storage_type
                ),
                "nominal_width": item.get("nominal_width"),
                "actual_width": item.get("actual_width"),
                "actual_width_meters": item.get("actual_width_meters"),
                "length_meters": item.get("length_meters"),
                "thickness": item.get("thickness"),
                "roll_no": item.get("roll_no"),
                "batch_no": item.get("batch_no"),
                "print_type": item.get("print_type"),
            }
        )
    elif item.get("category") == ROLL_PAPER_CATEGORY:
        query.update(
            {
                "storage_type": item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE,
                "thickness_micron": item.get("thickness_micron"),
                "width_meters": item.get("width_meters"),
                "length_meters": item.get("length_meters"),
            }
        )
    else:
        query.update(
            {
                "width": item.get("width"),
                "height": item.get("height"),
                "thickness": item.get("thickness"),
            }
        )
    return query


def get_inventory_sort():
    return [
        ("category", 1),
        ("brand", 1),
        ("type", 1),
        ("batch_roll_no", 1),
        ("roll_no", 1),
        ("batch_no", 1),
        ("width", 1),
        ("height", 1),
        ("thickness", 1),
    ]


def build_export_rows(items):
    rows = []
    for item in items:
        packaging = infer_item_packaging(item)
        breakdown = build_stock_breakdown(item) or {}
        export_brand = None if item.get("brand") == "__none__" else item.get("brand")
        export_type = None if item.get("type") == "__none__" else item.get("type")
        export_blanket_name = item.get("blanket_name") or item.get("brand")
        if item.get("category") == RUBBER_BLANKET_CATEGORY:
            export_blanket_name, _ = find_rubber_blanket_rule(export_blanket_name)
            export_blanket_name = export_blanket_name or item.get("blanket_name") or item.get("brand")
            export_brand = export_blanket_name
        row = {
            "Category": item["category"],
            "Brand": export_brand,
            "Type": export_type,
            "Blanket Name": export_blanket_name if item.get("category") == RUBBER_BLANKET_CATEGORY else None,
            "Storage Type": item.get("storage_type") if item.get("category") in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} else None,
            "Batch/Roll No": item.get("batch_roll_no"),
            "Width": item.get("width"),
            "Nominal Width": item.get("nominal_width") if item.get("category") == RUBBER_BLANKET_CATEGORY else None,
            "Actual Width": item.get("actual_width") if item.get("category") == RUBBER_BLANKET_CATEGORY else None,
            "Width Unit": item.get("width_unit"),
            "Length": item.get("height"),
            "Length Unit": item.get("length_unit"),
            "Size": f"{item.get('width')} X {item.get('height')}" if item.get("width") and item.get("height") else item.get("size"),
            "Thickness": item.get("thickness"),
            "Thickness Unit": item.get("thickness_unit"),
            "Thickness (Micron)": item.get("thickness_micron") if item.get("category") == ROLL_PAPER_CATEGORY else None,
            "Rolls": item.get("number_of_rolls"),
            "Number of Rolls": item.get("number_of_rolls"),
            "No. of Rolls": item.get("number_of_rolls") if item.get("category") in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} else None,
            "Number of Sheets": item.get("number_of_sheets") if item.get("category") in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} else None,
            "No. of Sheets": item.get("number_of_sheets") if item.get("category") in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} else None,
            "Roll No": item.get("roll_no"),
            "Batch No": item.get("batch_no"),
            "Print Type": item.get("print_type"),
            "Area per Roll": item.get("area_per_roll_sqm"),
            "Area per Sheet": item.get("area_per_sheet_sqm") if item.get("category") in {ROLL_PAPER_CATEGORY, RUBBER_BLANKET_CATEGORY} else None,
            "Quantity": item["quantity"],
            "Unit": item["unit"],
            "Product": export_brand if packaging and packaging["kind"] == "chemical" else None,
            "Pack Size": packaging.get("pack_size") if packaging else None,
            "Container Type": packaging.get("container_type") if packaging else None,
            "Containers/Box": (
                packaging.get("containers_per_box") or packaging.get("units_per_box")
                if packaging else None
            ),
            "Boxes": breakdown.get("boxes"),
            "Loose Units": breakdown.get("loose_units"),
            "Containers": breakdown.get("containers"),
            "Total Sheets": breakdown.get("total_sheets"),
        }
        rows.append(row)
    return rows


def get_excel_definition_for_item(item):
    category = item.get("category")
    packaging = infer_item_packaging(item)
    kind = packaging.get("kind") if packaging else None
    if category == RUBBER_BLANKET_CATEGORY and kind == "rubber_blankets":
        return EXCEL_SHEET_BY_KEY["rubber_blankets"]
    if category == ROLL_PAPER_CATEGORY:
        return EXCEL_SHEET_BY_KEY["calibrated_underpacking_paper"]
    if category == "Creasing Matrix":
        return EXCEL_SHEET_BY_KEY["creasing_matrix"]
    if kind == "chemical":
        return EXCEL_SHEET_BY_KEY["chemical"]
    if category == "CTCP Plates":
        return EXCEL_SHEET_BY_KEY["ctcp_plates"]
    return None


def excel_storage_type_label(value):
    return "Cut Piece" if value == M3Z_CUT_PIECE_STORAGE_TYPE else "Roll"


def excel_stock_unit_label(unit):
    if normalize_roll_stock_unit(unit):
        return "Sq.m"
    if normalize_sheet_stock_unit(unit):
        return "Sheets"
    if (clean_text(unit) or "").lower() == "ltr":
        return "L"
    return unit


def build_category_excel_row(item, definition):
    key = definition["key"]
    breakdown = build_stock_breakdown(item) or {}
    packaging = infer_item_packaging(item) or {}
    if key == "rubber_blankets":
        blanket_name, _ = find_rubber_blanket_rule(item.get("blanket_name") or item.get("brand"))
        return {
            "Blanket Name": blanket_name or item.get("blanket_name") or item.get("brand"),
            "Storage Type": excel_storage_type_label(item.get("storage_type")),
            "Thickness": item.get("thickness"),
            "Thickness Unit": item.get("thickness_unit") or "mm",
            "Print Type": item.get("print_type"),
            "Nominal Width": item.get("nominal_width") or item.get("width"),
            "Actual Width": item.get("actual_width") or item.get("width"),
            "Width Unit": item.get("width_unit") or "mm",
            "Length": item.get("length") or item.get("height"),
            "Length Unit": item.get("length_unit") or "m",
            "Number of Rolls": item.get("number_of_rolls"),
            "Number of Sheets": item.get("number_of_sheets"),
            "Area per Roll": item.get("area_per_roll_sqm"),
            "Area per Sheet": item.get("area_per_sheet_sqm"),
            "Quantity": item.get("quantity"),
            "Unit": excel_stock_unit_label(item.get("unit")),
        }
    if key == "calibrated_underpacking_paper":
        return {
            "Storage Type": excel_storage_type_label(item.get("storage_type")),
            "Thickness (Micron)": item.get("thickness_micron"),
            "Width": item.get("width"),
            "Width Unit": item.get("width_unit"),
            "Length": item.get("length") or item.get("height"),
            "Length Unit": item.get("length_unit"),
            "Number of Rolls": item.get("number_of_rolls"),
            "Number of Sheets": item.get("number_of_sheets"),
            "Area per Roll": item.get("area_per_roll_sqm"),
            "Area per Sheet": item.get("area_per_sheet_sqm"),
            "Quantity": item.get("quantity"),
            "Unit": excel_stock_unit_label(item.get("unit")),
        }
    if key == "creasing_matrix":
        return {
            "Thickness": item.get("thickness"),
            "Size": item.get("size") or build_size_label(item.get("width"), item.get("height")),
            "Boxes": breakdown.get("boxes"),
            "Loose Packets": breakdown.get("loose_units"),
            "Quantity (Pkt)": item.get("quantity"),
            "Unit": "pkt",
        }
    if key == "chemical":
        return {
            "Product": item.get("brand"),
            "Pack Size": packaging.get("pack_size"),
            "Container Type": packaging.get("container_type"),
            "Containers per Box": packaging.get("containers_per_box"),
            "Boxes": breakdown.get("boxes"),
            "Loose Containers": breakdown.get("loose_units"),
            "Total Containers": breakdown.get("containers"),
            "Total Quantity": item.get("quantity"),
            "Unit": excel_stock_unit_label(item.get("unit")),
        }
    if key == "ctcp_plates":
        return {
            "Thickness": item.get("thickness"),
            "Size": item.get("size") or build_size_label(item.get("width"), item.get("height")),
            "Boxes": breakdown.get("boxes", item.get("quantity")),
            "Sheets per Box": packaging.get("sheets_per_box", 50),
            "Total Sheets": breakdown.get("total_sheets"),
            "Quantity": item.get("quantity"),
            "Unit": "Box",
        }
    return {}


def build_excel_rows_by_sheet(items):
    rows_by_key = {definition["key"]: [] for definition in EXCEL_CATEGORY_SHEETS}
    legacy_items = []
    for item in items or []:
        definition = get_excel_definition_for_item(item)
        if not definition:
            legacy_items.append(item)
            continue
        rows_by_key[definition["key"]].append(build_category_excel_row(item, definition))
    return rows_by_key, legacy_items


def add_excel_defined_list(workbook, list_sheet, column_index, name, values):
    list_sheet.cell(row=1, column=column_index, value=name)
    for row_index, value in enumerate(values, start=2):
        list_sheet.cell(row=row_index, column=column_index, value=value)
    end_row = max(2, len(values) + 1)
    column_letter = get_column_letter(column_index)
    defined_name = DefinedName(
        name,
        attr_text=f"'{EXCEL_LISTS_SHEET}'!${column_letter}$2:${column_letter}${end_row}",
    )
    workbook.defined_names.add(defined_name)
    return column_index + 1


def build_excel_validation_lists(workbook):
    list_sheet = workbook.create_sheet(EXCEL_LISTS_SHEET)
    column_index = 1
    definitions = [
        ("StorageTypes", ["Roll", "Cut Piece"]),
        ("DimensionUnits", ["mm", "m", "inch"]),
        ("RubberBlanketNames", list(RUBBER_BLANKET_RULES.keys())),
        ("RubberThicknesses", sorted({value for rule in RUBBER_BLANKET_RULES.values() for value in rule.get("thickness_options", [])})),
        ("RubberPrintTypes", ["P", "W/O"]),
        ("M3ZThicknesses", sorted(M3Z_ALLOWED_THICKNESS_MICRONS, reverse=True)),
        ("CreasingThicknesses", [9, 11, 13]),
        ("ChemicalProducts", list(SPECIALIZED_CHEMICAL_PRODUCTS.keys())),
        ("ChemicalContainersPerBox", [4, 5, 12, 15, 18]),
        ("ChemicalContainerTypes", ["bottle"]),
        ("CTCPThicknesses", [0.30, 0.20]),
    ]
    for name, values in definitions:
        column_index = add_excel_defined_list(workbook, list_sheet, column_index, name, values)
    for thickness, sizes in CREASING_MATRIX_SIZES.items():
        column_index = add_excel_defined_list(workbook, list_sheet, column_index, f"Creasing_{thickness}", sizes)
    column_index = add_excel_defined_list(workbook, list_sheet, column_index, "CTCP_030", CTCP_PLATE_SIZES["0.30"])
    add_excel_defined_list(workbook, list_sheet, column_index, "CTCP_020", CTCP_PLATE_SIZES["0.20"])
    list_sheet.sheet_state = "veryHidden"


def add_excel_list_validation(sheet, column, formula, prompt):
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid value"
    validation.prompt = prompt
    validation.promptTitle = "Only Stock"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"{column}{EXCEL_DATA_START_ROW}:{column}{EXCEL_MAX_INPUT_ROW}")


def configure_category_sheet_validations(sheet, definition):
    columns = {header: get_column_letter(index) for index, header in enumerate(definition["headers"], start=1)}
    key = definition["key"]
    if key == "rubber_blankets":
        add_excel_list_validation(sheet, columns["Blanket Name"], "=RubberBlanketNames", "Choose an established blanket name.")
        add_excel_list_validation(sheet, columns["Storage Type"], "=StorageTypes", "Choose Roll or Cut Piece.")
        add_excel_list_validation(sheet, columns["Thickness"], "=RubberThicknesses", "The backend validates thickness for the selected blanket.")
        add_excel_list_validation(sheet, columns["Print Type"], "=RubberPrintTypes", "Use only when applicable to the selected blanket.")
        add_excel_list_validation(sheet, columns["Width Unit"], "=DimensionUnits", "Choose the width unit.")
        add_excel_list_validation(sheet, columns["Length Unit"], "=DimensionUnits", "Choose the length unit.")
        add_excel_list_validation(sheet, columns["Thickness Unit"], '"mm"', "Rubber Blanket thickness is stored in mm.")
    elif key == "calibrated_underpacking_paper":
        add_excel_list_validation(sheet, columns["Storage Type"], "=StorageTypes", "Choose Roll or Cut Piece.")
        add_excel_list_validation(sheet, columns["Thickness (Micron)"], "=M3ZThicknesses", "Choose a supported micron thickness.")
        add_excel_list_validation(sheet, columns["Width Unit"], "=DimensionUnits", "Choose the width unit.")
        add_excel_list_validation(sheet, columns["Length Unit"], "=DimensionUnits", "Choose the length unit.")
    elif key == "creasing_matrix":
        add_excel_list_validation(sheet, columns["Thickness"], "=CreasingThicknesses", "Choose 9, 11, or 13 mm.")
        size_column = columns["Size"]
        thickness_column = columns["Thickness"]
        add_excel_list_validation(
            sheet,
            size_column,
            f'=INDIRECT("Creasing_"&TEXT(${thickness_column}{EXCEL_DATA_START_ROW},"0"))',
            "The size list changes with thickness.",
        )
    elif key == "chemical":
        add_excel_list_validation(sheet, columns["Product"], "=ChemicalProducts", "Choose an established chemical product.")
        add_excel_list_validation(sheet, columns["Container Type"], "=ChemicalContainerTypes", "Choose the container type.")
        add_excel_list_validation(sheet, columns["Containers per Box"], "=ChemicalContainersPerBox", "The backend validates the product configuration.")
    elif key == "ctcp_plates":
        add_excel_list_validation(sheet, columns["Thickness"], "=CTCPThicknesses", "Choose 0.30 or 0.20.")
        size_column = columns["Size"]
        thickness_column = columns["Thickness"]
        add_excel_list_validation(
            sheet,
            size_column,
            f'=IF(TEXT(${thickness_column}{EXCEL_DATA_START_ROW},"0.00")="0.30",CTCP_030,CTCP_020)',
            "The size list changes with thickness.",
        )


def style_readme_sheet(sheet, workbook_title):
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = workbook_title
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 36
    sheet.merge_cells("A3:F3")
    sheet["A3"] = "This workbook contains separate sheets for each inventory item category. Enter data only in the relevant sheet."
    sheet["A3"].font = Font(name="Aptos", size=11, color="334155")
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 32

    instructions = [
        "Do not change column headers.",
        "Enter data only in the correct category sheet; empty rows are ignored.",
        "Required fields are marked in red and controlled fields include dropdowns.",
        "The backend recalculates all derived quantities and remains authoritative.",
        "Roll stock is normalized to Sq.m; cut pieces are stored in Sheets.",
        "Update mode is complete only for non-empty category sheets supplied in the workbook.",
    ]
    sheet["A5"] = "Instructions"
    sheet["A5"].font = Font(name="Aptos", size=12, bold=True, color="0F172A")
    for index, instruction in enumerate(instructions, start=6):
        sheet.merge_cells(start_row=index, start_column=1, end_row=index, end_column=6)
        sheet.cell(index, 1, f"• {instruction}")
        sheet.cell(index, 1).font = Font(name="Aptos", size=10, color="475569")

    table_row = 13
    sheet.cell(table_row, 1, "Sheet Name")
    sheet.cell(table_row, 2, "Item Category")
    for cell in sheet[table_row][0:2]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    category_rows = [
        ("01_Rubber_Blankets", "Rubber Blankets — Roll / Cut Piece"),
        ("02_Calibrated_Underpacking", "05 - Calibrated Underpacking Paper — Roll / Cut Piece"),
        ("03_Creasing_Matrix", "Creasing Matrix — Packets"),
        ("04_Chemical", "Chemical — Litre / kg"),
        ("05_CTCP_Plates", "CTCP Plates — Boxes"),
    ]
    thin_border = Border(bottom=Side(style="thin", color="D8E2DC"))
    for row_index, values in enumerate(category_rows, start=table_row + 1):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.font = Font(name="Aptos", size=10, color="334155")
            cell.border = thin_border
            cell.fill = PatternFill("solid", fgColor="F8FAFC" if row_index % 2 == 0 else "FFFFFF")

    notes_row = table_row + len(category_rows) + 2
    sheet[notes_row][0].value = "Important stock rules"
    sheet[notes_row][0].font = Font(name="Aptos", size=12, bold=True, color="0F172A")
    notes = [
        "Creasing Matrix: 10 packets = 1 box; inventory quantity is packets.",
        "Chemical: boxes and loose containers are normalized to litre/kg quantity.",
        "CTCP Plates: 1 box = 50 sheets; inventory quantity remains boxes.",
        "Unknown worksheets are never interpreted as inventory.",
    ]
    for index, note in enumerate(notes, start=notes_row + 1):
        sheet.merge_cells(start_row=index, start_column=1, end_row=index, end_column=6)
        sheet.cell(index, 1, f"• {note}")
        sheet.cell(index, 1).font = Font(name="Aptos", size=10, color="475569")
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 54
    for column in "CDEF":
        sheet.column_dimensions[column].width = 12
    sheet.freeze_panes = "A5"


def style_category_sheet(sheet, definition, rows, workbook_mode):
    headers = definition["headers"]
    last_column = get_column_letter(len(headers))
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = f"ONLY STOCK — {definition['label'].upper()}"
    sheet["A1"].fill = PatternFill("solid", fgColor=definition["accent"])
    sheet["A1"].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    mode_note = "Existing values may be edited before Update." if workbook_mode == "update" else "Use one row per inventory identity."
    sheet["A2"] = f"{definition['instruction']} {mode_note}"
    sheet["A2"].fill = PatternFill("solid", fgColor="F8FAFC")
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color="475569")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34

    required_headers = set(definition["required_headers"])
    derived_headers = set(definition["derived_headers"])
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(EXCEL_HEADER_ROW, column_index, header)
        if header in derived_headers:
            fill_color = "64748B"
            cell.comment = Comment("Calculated or normalized by Only Stock. Do not use this value to override physical inputs.", "Only Stock")
        elif header in required_headers:
            fill_color = "B91C1C"
        else:
            fill_color = "0F766E"
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[EXCEL_HEADER_ROW].height = 38

    for row_offset, row in enumerate(rows, start=EXCEL_DATA_START_ROW):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row_offset, column_index, row.get(header))
            cell.font = Font(name="Aptos", size=10, color="334155")
            cell.alignment = Alignment(vertical="center")
            if header in derived_headers:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")
            elif row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAF9")
        sheet.row_dimensions[row_offset].height = 22

    numeric_headers = {
        "Thickness", "Thickness (Micron)", "Nominal Width", "Actual Width", "Width", "Length",
        "Number of Rolls", "Number of Sheets", "Area per Roll", "Area per Sheet", "Quantity",
        "Boxes", "Loose Packets", "Quantity (Pkt)", "Pack Size", "Containers per Box",
        "Loose Containers", "Total Containers", "Total Quantity", "Sheets per Box", "Total Sheets",
    }
    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        longest = max([len(str(header)), *[len(str(row.get(header) or "")) for row in rows]] or [len(header)])
        sheet.column_dimensions[column_letter].width = min(max(longest + 3, 13), 28)
        if header in numeric_headers:
            sheet[f"{column_letter}{EXCEL_DATA_START_ROW}:{column_letter}{max(EXCEL_DATA_START_ROW, EXCEL_DATA_START_ROW + len(rows) - 1)}"][0][0].number_format = "0.####"
    sheet.freeze_panes = f"A{EXCEL_DATA_START_ROW}"
    sheet.auto_filter.ref = f"A{EXCEL_HEADER_ROW}:{last_column}{max(EXCEL_HEADER_ROW, EXCEL_HEADER_ROW + len(rows))}"


def build_inventory_workbook(items=None, workbook_mode="template"):
    # Reload the shared catalog when a workbook is requested so Excel always
    # reflects the current JSON file, even when the Flask process stayed up.
    global RUBBER_BLANKET_RULES
    RUBBER_BLANKET_RULES = load_rubber_blanket_rules()
    workbook = Workbook()
    readme = workbook.active
    readme.title = EXCEL_README_SHEET
    workbook_title = {
        "template": "ONLY STOCK — EXCEL IMPORT TEMPLATE",
        "current": "ONLY STOCK — CURRENT INVENTORY",
        "update": "ONLY STOCK — INVENTORY UPDATE WORKBOOK",
    }.get(workbook_mode, "ONLY STOCK — INVENTORY WORKBOOK")
    style_readme_sheet(readme, workbook_title)

    rows_by_key, legacy_items = build_excel_rows_by_sheet(items or [])
    category_sheets = {}
    for definition in EXCEL_CATEGORY_SHEETS:
        sheet = workbook.create_sheet(definition["sheet_name"])
        rows = rows_by_key[definition["key"]]
        style_category_sheet(sheet, definition, rows, workbook_mode)
        category_sheets[definition["key"]] = sheet

    if legacy_items:
        legacy_sheet = workbook.create_sheet(EXCEL_LEGACY_SHEET)
        legacy_rows = build_export_rows(legacy_items)
        legacy_definition = {
            "key": "legacy",
            "label": "Legacy Inventory",
            "accent": "475569",
            "headers": EXCEL_COLUMNS,
            "required_headers": REQUIRED_EXCEL_COLUMNS,
            "derived_headers": [],
            "instruction": "Compatibility sheet for inventory categories that still use the original generic Excel structure.",
        }
        style_category_sheet(legacy_sheet, legacy_definition, legacy_rows, workbook_mode)

    build_excel_validation_lists(workbook)
    for definition in EXCEL_CATEGORY_SHEETS:
        configure_category_sheet_validations(category_sheets[definition["key"]], definition)
    workbook.active = 0
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def send_inventory_workbook(items, download_name, workbook_mode):
    return send_file(
        build_inventory_workbook(items, workbook_mode),
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def excel_json_value(value):
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, datetime):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass
    return value


def excel_value_or(value, default):
    return value if parse_optional_text(value) is not None else default


def normalize_excel_header(value):
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    header = re.sub(r"\s+", " ", str(value)).strip()
    return header or None


def locate_excel_header_row(excel_file, sheet_name, expected_headers):
    preview = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=12)
    expected = set(expected_headers)
    best_index = None
    best_headers = set()
    for row_index, values in preview.iterrows():
        headers = {normalize_excel_header(value) for value in values.tolist()}
        headers.discard(None)
        if len(headers & expected) > len(best_headers & expected):
            best_index = int(row_index)
            best_headers = headers
    return best_index, best_headers


def read_excel_sheet_records(excel_file, sheet_name, definition):
    required_headers = definition["required_headers"]
    header_index, detected_headers = locate_excel_header_row(excel_file, sheet_name, definition["headers"])
    missing_headers = [header for header in required_headers if header not in detected_headers]
    if header_index is None or missing_headers:
        return [], {
            "sheet": sheet_name,
            "field": "Headers",
            "problem": (
                f"{sheet_name} is missing required column: {missing_headers[0]}"
                if missing_headers
                else f"{sheet_name} does not contain a recognizable header row"
            ),
            "missing_headers": missing_headers,
            "expected_headers": definition["headers"],
        }, header_index

    dataframe = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_index)
    normalized_columns = [normalize_excel_header(column) or f"Unnamed_{index}" for index, column in enumerate(dataframe.columns)]
    dataframe.columns = normalized_columns
    available_headers = [header for header in definition["headers"] if header in dataframe.columns]
    records = []
    for offset, row in dataframe.iterrows():
        values = {header: row.get(header) for header in available_headers}
        if not any(parse_optional_text(value) is not None for value in values.values()):
            continue
        records.append(
            {
                "excel_row": int(header_index) + 2 + int(offset),
                "values": values,
            }
        )
    return records, None, header_index


def parse_category_excel_row(definition, row):
    key = definition["key"]
    if key == "rubber_blankets":
        return build_item_payload(
            {
                "category": RUBBER_BLANKET_CATEGORY,
                "blanket_name": row.get("Blanket Name"),
                "storage_type": row.get("Storage Type"),
                "thickness": row.get("Thickness"),
                "thickness_unit": excel_value_or(row.get("Thickness Unit"), "mm"),
                "print_type": row.get("Print Type"),
                "nominal_width": row.get("Nominal Width"),
                "actual_width": row.get("Actual Width"),
                "width_unit": row.get("Width Unit"),
                "length": row.get("Length"),
                "length_unit": row.get("Length Unit"),
                "number_of_rolls": row.get("Number of Rolls"),
                "number_of_sheets": row.get("Number of Sheets"),
            }
        )
    if key == "calibrated_underpacking_paper":
        return build_item_payload(
            {
                "category": ROLL_PAPER_CATEGORY,
                "storage_type": row.get("Storage Type"),
                "thickness_micron": row.get("Thickness (Micron)"),
                "thickness_unit": "micron",
                "width": row.get("Width"),
                "width_unit": row.get("Width Unit"),
                "length": row.get("Length"),
                "length_unit": row.get("Length Unit"),
                "number_of_rolls": row.get("Number of Rolls"),
                "number_of_sheets": row.get("Number of Sheets"),
            }
        )
    if key == "creasing_matrix":
        return build_item_payload(
            {
                "category": "Creasing Matrix",
                "thickness": row.get("Thickness"),
                "size": row.get("Size"),
                "boxes": row.get("Boxes"),
                "loose_units": row.get("Loose Packets"),
                "quantity": row.get("Quantity (Pkt)"),
                "unit": excel_value_or(row.get("Unit"), "pkt"),
            }
        )
    if key == "chemical":
        product_name, product_config = find_chemical_product(row.get("Product"))
        if not product_config:
            return None, "product must be selected from the established Chemical product list"
        return build_item_payload(
            {
                "category": product_config["category"],
                "product": product_name,
                "pack_size": row.get("Pack Size"),
                "container_type": row.get("Container Type"),
                "containers_per_box": row.get("Containers per Box"),
                "boxes": row.get("Boxes"),
                "loose_units": row.get("Loose Containers"),
                "containers": row.get("Total Containers"),
                "quantity": row.get("Total Quantity"),
                "unit": excel_value_or(row.get("Unit"), product_config["unit"]),
            }
        )
    if key == "ctcp_plates":
        return build_item_payload(
            {
                "category": "CTCP Plates",
                "thickness": row.get("Thickness"),
                "size": row.get("Size"),
                "boxes": row.get("Boxes"),
                "total_sheets": row.get("Total Sheets"),
                "quantity": row.get("Quantity"),
                "unit": excel_value_or(row.get("Unit"), "box"),
            }
        )
    return None, "unsupported category sheet"


def infer_excel_error_field(problem):
    lowered = (problem or "").lower()
    field_names = [
        "blanket name", "storage type", "thickness", "actual width", "nominal width", "width",
        "length", "number of rolls", "number of sheets", "print type", "size", "product",
        "containers per box", "loose containers", "container", "boxes", "total sheets", "quantity", "unit",
    ]
    for field_name in field_names:
        if field_name in lowered:
            return field_name.title()
    return "Row"


def parse_legacy_excel_sheet(excel_file, sheet_name):
    legacy_definition = {
        "key": "legacy",
        "sheet_name": sheet_name,
        "label": "Legacy Inventory",
        "headers": EXCEL_COLUMNS,
        "required_headers": REQUIRED_EXCEL_COLUMNS,
        "derived_headers": [],
    }
    records, sheet_error, _ = read_excel_sheet_records(excel_file, sheet_name, legacy_definition)
    if sheet_error:
        return None, sheet_error
    parsed_rows = []
    for record in records:
        item, error = process_excel_row(record["values"])
        parsed_rows.append({**record, "item": item, "error": error})
    return {
        "definition": legacy_definition,
        "sheet_name": sheet_name,
        "rows": parsed_rows,
        "sheet_errors": [],
    }, None


def parse_uploaded_excel_workbook(uploaded_file):
    filename = (uploaded_file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        return None, "Only .xlsx and .xls Excel workbooks are supported"
    file_bytes = uploaded_file.read()
    if not file_bytes:
        return None, "Excel file is empty"
    try:
        excel_file = pd.ExcelFile(BytesIO(file_bytes))
    except Exception:
        return None, "Unable to read Excel file"

    sheet_names = list(excel_file.sheet_names)
    recognized_names = [name for name in sheet_names if name in EXCEL_SHEET_BY_NAME]
    parsed_sheets = []
    warnings = []
    workbook_type = "multi_sheet" if recognized_names else "legacy"

    if recognized_names:
        for sheet_name in sheet_names:
            if sheet_name in {EXCEL_README_SHEET, EXCEL_LISTS_SHEET}:
                continue
            definition = EXCEL_SHEET_BY_NAME.get(sheet_name)
            if not definition:
                if sheet_name == EXCEL_LEGACY_SHEET:
                    legacy_sheet, legacy_error = parse_legacy_excel_sheet(excel_file, sheet_name)
                    if legacy_error:
                        parsed_sheets.append(
                            {
                                "definition": {
                                    "key": "legacy", "sheet_name": sheet_name, "label": "Legacy Inventory",
                                    "headers": EXCEL_COLUMNS, "required_headers": REQUIRED_EXCEL_COLUMNS,
                                },
                                "sheet_name": sheet_name,
                                "rows": [],
                                "sheet_errors": [legacy_error],
                            }
                        )
                    else:
                        parsed_sheets.append(legacy_sheet)
                    continue
                warnings.append(
                    {
                        "sheet": sheet_name,
                        "problem": "Unsupported worksheet was ignored.",
                        "suggestion": "Use one of the documented category sheet names.",
                    }
                )
                continue
            records, sheet_error, _ = read_excel_sheet_records(excel_file, sheet_name, definition)
            parsed_rows = []
            if not sheet_error:
                for record in records:
                    item, error = parse_category_excel_row(definition, record["values"])
                    parsed_rows.append({**record, "item": item, "error": error})
            parsed_sheets.append(
                {
                    "definition": definition,
                    "sheet_name": sheet_name,
                    "rows": parsed_rows,
                    "sheet_errors": [sheet_error] if sheet_error else [],
                }
            )
    else:
        legacy_sheet = None
        legacy_error = None
        for sheet_name in sheet_names:
            if sheet_name in {EXCEL_README_SHEET, EXCEL_LISTS_SHEET}:
                continue
            candidate, candidate_error = parse_legacy_excel_sheet(excel_file, sheet_name)
            if candidate is not None:
                legacy_sheet = candidate
                break
            legacy_error = candidate_error
        if legacy_sheet is None:
            return None, "Unsupported Excel workbook structure. Expected category sheets or a legacy sheet containing Category headers."
        parsed_sheets.append(legacy_sheet)
        if len(sheet_names) > 1:
            warnings.append(
                {
                    "sheet": legacy_sheet["sheet_name"],
                    "problem": "Legacy single-sheet workbook detected.",
                    "suggestion": "Use the new category-specific template for the clearest validation and preview.",
                }
            )

    seen_keys = {}
    for sheet in parsed_sheets:
        for row in sheet["rows"]:
            if row.get("error") or not row.get("item"):
                continue
            item_key = build_item_key(row["item"])
            if item_key in seen_keys:
                previous = seen_keys[item_key]
                row["error"] = f"duplicate item; first entered in {previous['sheet']} row {previous['row']}"
                row["item"] = None
            else:
                seen_keys[item_key] = {"sheet": sheet["sheet_name"], "row": row["excel_row"]}

    return {
        "workbook_type": workbook_type,
        "sheet_names": sheet_names,
        "sheets": parsed_sheets,
        "warnings": warnings,
    }, None


def get_parsed_excel_items(parsed_workbook):
    return [
        row["item"]
        for sheet in parsed_workbook["sheets"]
        for row in sheet["rows"]
        if row.get("item") is not None and not row.get("error")
    ]


def get_excel_update_scopes(parsed_workbook):
    scopes = []
    for sheet in parsed_workbook["sheets"]:
        if not sheet["rows"] or sheet["sheet_errors"]:
            continue
        key = sheet["definition"]["key"]
        if key == "legacy":
            categories = sorted(
                {row["item"]["category"] for row in sheet["rows"] if row.get("item") and not row.get("error")}
            )
            if categories:
                scopes.append({"key": "legacy", "label": "Legacy Inventory", "categories": categories})
        else:
            scopes.append({"key": key, "label": sheet["definition"]["label"]})
    return scopes


def get_excel_scope_query(scope):
    key = scope["key"]
    if key == "rubber_blankets":
        return {
            "category": RUBBER_BLANKET_CATEGORY,
            "blanket_name": {"$exists": True, "$ne": None},
        }
    if key == "calibrated_underpacking_paper":
        return {"category": ROLL_PAPER_CATEGORY}
    if key == "creasing_matrix":
        return {"category": "Creasing Matrix"}
    if key == "chemical":
        return {
            "category": {"$in": sorted(CHEMICAL_CATEGORIES)},
            "brand": {"$in": sorted(SPECIALIZED_CHEMICAL_PRODUCTS.keys())},
        }
    if key == "ctcp_plates":
        return {"category": "CTCP Plates"}
    if key == "legacy":
        return {"category": {"$in": scope.get("categories", [])}}
    return {"_id": {"$exists": False}}


def find_excel_update_deletions(parsed_workbook, inventory_collection, seen_keys):
    deletions = []
    found_ids = set()
    for scope in get_excel_update_scopes(parsed_workbook):
        for existing_item in inventory_collection.find(get_excel_scope_query(scope)):
            item_id = str(existing_item.get("_id"))
            if item_id in found_ids or build_item_key(existing_item) in seen_keys:
                continue
            found_ids.add(item_id)
            deletions.append(existing_item)
    return deletions


def build_excel_preview_response(parsed_workbook, inventory_collection, upload_mode):
    preview_sheets = []
    all_errors = []
    valid_items = []
    seen_keys = set()
    source_by_key = {}
    for sheet in parsed_workbook["sheets"]:
        definition = sheet["definition"]
        preview_rows = []
        for sheet_error in sheet["sheet_errors"]:
            all_errors.append(sheet_error)
        for row in sheet["rows"]:
            item = row.get("item")
            problem = row.get("error")
            if item and not problem:
                valid_items.append(item)
                item_key = build_item_key(item)
                seen_keys.add(item_key)
                source_by_key[item_key] = {"sheet": sheet["sheet_name"], "row": row["excel_row"]}
                values = (
                    build_category_excel_row(item, definition)
                    if definition["key"] != "legacy"
                    else build_export_rows([item])[0]
                )
                status = "valid"
                errors = []
            else:
                values = row["values"]
                status = "error"
                error_detail = {
                    "sheet": sheet["sheet_name"],
                    "row": row["excel_row"],
                    "field": infer_excel_error_field(problem),
                    "problem": problem,
                    "suggestion": "Correct the value using the sheet instructions and dropdown options.",
                }
                all_errors.append(error_detail)
                errors = [error_detail]
            preview_rows.append(
                {
                    "excel_row": row["excel_row"],
                    "status": status,
                    "values": {header: excel_json_value(values.get(header)) for header in definition["headers"]},
                    "errors": errors,
                }
            )
        preview_sheets.append(
            {
                "key": definition["key"],
                "sheet_name": sheet["sheet_name"],
                "label": definition["label"],
                "columns": definition["headers"],
                "row_count": len(sheet["rows"]),
                "valid_count": sum(1 for row in preview_rows if row["status"] == "valid"),
                "error_count": sum(1 for row in preview_rows if row["status"] == "error") + len(sheet["sheet_errors"]),
                "rows": preview_rows,
                "sheet_errors": sheet["sheet_errors"],
            }
        )

    to_add = 0
    to_update = 0
    unchanged = 0
    for item in valid_items:
        existing_item = inventory_collection.find_one(get_item_identity_query(item))
        if not existing_item:
            to_add += 1
        elif get_changed_fields(existing_item, item):
            to_update += 1
        else:
            unchanged += 1
    deletions = (
        find_excel_update_deletions(parsed_workbook, inventory_collection, seen_keys)
        if upload_mode == "update" and not all_errors
        else []
    )
    scopes = get_excel_update_scopes(parsed_workbook)
    return {
        "message": "Workbook parsed successfully",
        "mode": upload_mode,
        "workbook_type": parsed_workbook["workbook_type"],
        "sheet_names": parsed_workbook["sheet_names"],
        "sheets": preview_sheets,
        "total_rows": sum(sheet["row_count"] for sheet in preview_sheets),
        "validation": {
            "valid_rows": len(valid_items),
            "warning_count": len(parsed_workbook["warnings"]),
            "error_rows": len(all_errors),
            "warnings": parsed_workbook["warnings"],
            "errors": all_errors,
        },
        "update_summary": {
            "add": to_add,
            "update": to_update,
            "delete": len(deletions),
            "unchanged": unchanged,
            "scope": scopes,
        },
        "can_apply": bool(valid_items) and not all_errors,
    }


EXCEL_MUTABLE_FIELDS = [
    "quantity", "unit", "batch_roll_no", "size", "width", "height", "width_unit", "length",
    "length_unit", "thickness", "thickness_unit", "thickness_micron", "storage_type", "number_of_rolls",
    "number_of_sheets", "width_meters", "length_meters", "area_per_roll_sqm", "area_per_sheet_sqm",
    "blanket_name", "nominal_width", "actual_width", "actual_width_meters", "roll_no", "batch_no",
    "print_type", "packaging", "format_size", "format_unit",
]


def apply_parsed_excel_workbook(parsed_workbook, inventory_collection, upload_mode, reason):
    inserted = 0
    updated = 0
    unchanged = 0
    seen_keys = set()
    source_by_key = {}
    for sheet in parsed_workbook["sheets"]:
        for row in sheet["rows"]:
            item = row.get("item")
            if not item or row.get("error"):
                continue
            item_key = build_item_key(item)
            seen_keys.add(item_key)
            source_by_key[item_key] = {"sheet": sheet["sheet_name"], "row": row["excel_row"]}
            existing_item = inventory_collection.find_one(get_item_identity_query(item))
            details = {"mode": upload_mode, **source_by_key[item_key]}
            if existing_item:
                changes = get_changed_fields(existing_item, item)
                if not changes:
                    unchanged += 1
                    continue
                updates = {field: item.get(field) for field in EXCEL_MUTABLE_FIELDS}
                updates["updated_at"] = now_ist()
                inventory_collection.update_one({"_id": existing_item["_id"]}, {"$set": updates})
                latest_item = inventory_collection.find_one({"_id": existing_item["_id"]})
                log_stock_change(
                    latest_item,
                    "excel_update",
                    existing_item["quantity"],
                    latest_item["quantity"],
                    "excel",
                    reason,
                    {**details, "changes": changes},
                )
                updated += 1
            else:
                result = inventory_collection.insert_one(item)
                item["_id"] = result.inserted_id
                log_stock_change(item, "excel_create", 0, item["quantity"], "excel", reason, details)
                inserted += 1

    deleted = 0
    if upload_mode == "update":
        for existing_item in find_excel_update_deletions(parsed_workbook, inventory_collection, seen_keys):
            inventory_collection.delete_one({"_id": existing_item["_id"]})
            log_stock_change(
                existing_item,
                "excel_delete",
                existing_item["quantity"],
                0,
                "excel",
                reason,
                {"mode": upload_mode, "delete_source": "missing_from_category_sheet"},
            )
            deleted += 1
    return {
        "message": "Excel processed successfully",
        "inserted": inserted,
        "updated": updated,
        "deleted": deleted,
        "unchanged": unchanged,
        "total_rows": len(get_parsed_excel_items(parsed_workbook)),
        "mode": upload_mode,
    }


def normalized_adjustment_date(date_text):
    try:
        parsed_date = datetime.strptime(date_text, "%Y-%m-%d")
    except (TypeError, ValueError):
        parsed_date = now_ist().replace(tzinfo=None)
    return parsed_date.strftime("%Y-%m-%d")


def split_adjustment_manufacturer_and_unit(value, fallback_unit):
    text = clean_adjustment_part(value)
    match = re.match(r"^(.*?)(?:\s+(Mtr|sq\.mtr|square\.mtr|sqmtr|pcs))?$", text, flags=re.IGNORECASE)
    if not match:
        return text, fallback_unit
    manufacturer_batch = clean_adjustment_part(match.group(1))
    usage_unit = normalize_adjustment_unit(match.group(2) or fallback_unit)
    return manufacturer_batch, usage_unit


def parse_adjustment_number(value):
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def parse_adjustment_quantity_and_unit(value, fallback_unit):
    text = clean_adjustment_part(value)
    quantity = parse_adjustment_number(text)
    unit_match = re.search(r"\b(Mtr|sq\.mtr|square\.mtr|sqmtr|pcs)\b", text, flags=re.IGNORECASE)
    usage_unit = normalize_adjustment_unit(unit_match.group(1) if unit_match else fallback_unit)
    return quantity, usage_unit


def clean_adjustment_part(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_adjustment_unit(value):
    text = clean_adjustment_part(value).lower()
    if text in {"sq.mtr", "square.mtr", "square mtr", "sqmtr", "sq mtr"}:
        return "sq.mtr"
    if text in {"mtr", "meter", "meters"}:
        return "Mtr"
    if text in {"pcs", "piece", "pieces"}:
        return "pcs"
    return clean_adjustment_part(value) or "Mtr"


def normalize_adjustment_item_name(value):
    text = clean_adjustment_part(value)
    if not text:
        return ""

    width_match = re.match(r"^(?P<width>\d+(?:\.\d+)?)\s*mm\b\s*(?P<rest>.+)$", text, flags=re.IGNORECASE)
    thickness_match = re.search(r"(?P<thickness>\d+(?:\.\d+)?)\s*mm$", text, flags=re.IGNORECASE)

    width = None
    thickness = None
    core_text = text

    if width_match:
        width = width_match.group("width")
        core_text = clean_adjustment_part(width_match.group("rest"))

    if thickness_match:
        thickness = thickness_match.group("thickness")
        core_text = clean_adjustment_part(re.sub(r"(?P<thickness>\d+(?:\.\d+)?)\s*mm$", "", core_text, flags=re.IGNORECASE))

    if width and thickness:
        return f"{core_text} - {width}mm {thickness}mm"
    if width:
        return f"{core_text} - {width}mm"
    return text


def parse_inventory_adjustment_text(text, defaults):
    rows = []
    current_item = DEFAULT_ADJUSTMENT_ITEM_NAME
    for raw_line in str(text or "").splitlines():
        line = clean_adjustment_part(raw_line)
        if not line:
            continue

        copied_row_match = re.match(
            r"^(?P<batch>[A-Za-z]{1,3}\d+)\s*-\s*"
            r"(?P<mfg>-?\d+(?:\.\d+)?)\s+(?P<mfg_unit>[A-Za-z.]+)\s+"
            r"(?P<warehouse>Main\s+Location)\s+"
            r"(?P<quantity>-?\d+(?:\.\d+)?)\s+(?P<quantity_unit>[A-Za-z.]+)\s+"
            r"(?P<cost>-?\d+(?:\.\d+)?)",
            line,
            flags=re.IGNORECASE,
        )
        if copied_row_match:
            quantity = parse_adjustment_number(copied_row_match.group("quantity"))
            cost_price = parse_adjustment_number(copied_row_match.group("cost"))
            manufacturer_unit = normalize_adjustment_unit(copied_row_match.group("mfg_unit"))
            usage_unit = normalize_adjustment_unit(copied_row_match.group("quantity_unit") or copied_row_match.group("mfg_unit"))
            rows.append(
                {
                    "Reference#": "app",
                    "Batch Reference#": copied_row_match.group("batch"),
                    "InventoryAdjustment#": "",
                    "Date": defaults["date_display"],
                    "Status": "Adjusted",
                    "InventoryAdjustment ID": "",
                    "Adjustment Type": "quantity",
                    "Reason": defaults["reason"],
                    "Item Name": current_item,
                    "Item ID": "",
                    "Batch Number": copied_row_match.group("batch"),
                    "Manufacturer Batch#": f"{copied_row_match.group('mfg')} {manufacturer_unit}",
                    "Warehouse Name": copied_row_match.group("warehouse"),
                    "Account": "Cost of Goods Sold",
                    "Inventory Account": "Inventory Asset",
                    "Quantity Adjusted": quantity,
                    "Cost Price": cost_price,
                    "Value Adjusted": "",
                    "Usage unit": usage_unit,
                }
            )
            continue

        detail_line = re.sub(r"^\d+\.\s*", "", line).strip()
        parts = [clean_adjustment_part(part) for part in re.split(r"\s+-\s+", detail_line) if clean_adjustment_part(part)]
        if len(parts) >= 4 and re.match(r"^[A-Za-z]{1,3}\d+$", parts[0]):
            batch_reference = parts[0]
            manufacturer_batch = parts[1]
            if len(parts) >= 5:
                warehouse_name = parts[2] or DEFAULT_ADJUSTMENT_WAREHOUSE
                quantity, usage_unit = parse_adjustment_quantity_and_unit(parts[3], defaults["unit"])
                cost_price = parse_adjustment_number(parts[4])
            else:
                warehouse_name = DEFAULT_ADJUSTMENT_WAREHOUSE
                quantity, usage_unit = parse_adjustment_quantity_and_unit(parts[2], defaults["unit"])
                cost_price = parse_adjustment_number(parts[3])
            rows.append(
                {
                    "Reference#": "app",
                    "Batch Reference#": batch_reference,
                    "InventoryAdjustment#": "",
                    "Date": defaults["date_display"],
                    "Status": "Adjusted",
                    "InventoryAdjustment ID": "",
                    "Adjustment Type": "quantity",
                    "Reason": defaults["reason"],
                    "Item Name": current_item,
                    "Item ID": "",
                    "Batch Number": batch_reference,
                    "Manufacturer Batch#": manufacturer_batch,
                    "Warehouse Name": warehouse_name,
                    "Account": "Cost of Goods Sold",
                    "Inventory Account": "Inventory Asset",
                    "Quantity Adjusted": quantity,
                    "Cost Price": cost_price,
                    "Value Adjusted": "",
                    "Usage unit": usage_unit,
                }
            )
            continue

        current_item = normalize_adjustment_item_name(re.sub(r"^\d+\.\s*", "", line).strip()) or DEFAULT_ADJUSTMENT_ITEM_NAME

    return rows


def extract_adjustment_text_from_xls(uploaded_file):
    book = xlrd.open_workbook(file_contents=uploaded_file.read())
    lines = []
    for sheet in book.sheets():
        if sheet.nrows:
            headers = [clean_adjustment_part(sheet.cell_value(0, column_index)) for column_index in range(sheet.ncols)]
            header_map = {header: index for index, header in enumerate(headers) if header}
            required_headers = {"Item Name", "Quantity Adjusted", "Cost Price"}
            batch_header = "Batch Reference#" if "Batch Reference#" in header_map else "Reference#"
            if required_headers.issubset(header_map) and batch_header in header_map:
                current_item = ""
                for row_index in range(1, sheet.nrows):
                    def cell(header):
                        column_index = header_map.get(header)
                        return clean_adjustment_part(sheet.cell_value(row_index, column_index)) if column_index is not None else ""

                    item_name = cell("Item Name") or DEFAULT_ADJUSTMENT_ITEM_NAME
                    batch_reference = cell(batch_header)
                    quantity = cell("Quantity Adjusted")
                    cost_price = cell("Cost Price")
                    usage_unit = normalize_adjustment_unit(cell("Usage unit") or "Mtr")
                    manufacturer_batch = cell("Manufacturer Batch#") or cell("Description") or cell("Item Desc")
                    warehouse_name = cell("Warehouse Name") or DEFAULT_ADJUSTMENT_WAREHOUSE
                    if not re.match(r"^[A-Za-z]{1,3}\d+$", batch_reference):
                        continue
                    if item_name != current_item:
                        lines.append(item_name)
                        current_item = item_name
                    lines.append(f"{batch_reference} - {manufacturer_batch} - {warehouse_name} - {quantity} {usage_unit} - {cost_price}")
                continue

        for row_index in range(sheet.nrows):
            values = []
            for column_index in range(sheet.ncols):
                value = sheet.cell_value(row_index, column_index)
                if value not in {"", None}:
                    values.append(clean_adjustment_part(value))
            if values:
                lines.append(" - ".join(values) if len(values) > 1 else values[0])
    return "\n".join(lines)


def send_inventory_adjustment_xls(rows):
    output = BytesIO()
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Inventory Adjustment")
    header_style = xlwt.easyxf("font: bold on; align: horiz center")
    number_style = xlwt.easyxf(num_format_str="0.00")
    price_style = xlwt.easyxf(num_format_str="0.000000")

    for column_index, column in enumerate(INVENTORY_ADJUSTMENT_COLUMNS):
        sheet.write(0, column_index, column, header_style)
        sheet.col(column_index).width = max(3500, len(column) * 320)

    for row_index, row in enumerate(rows, start=1):
        for column_index, column in enumerate(INVENTORY_ADJUSTMENT_COLUMNS):
            value = row.get(column, "")
            style = xlwt.Style.default_style
            if column == "Quantity Adjusted" and value not in {"", None}:
                style = number_style
            elif column in {"Cost Price", "Value Adjusted"} and value not in {"", None}:
                style = price_style
            sheet.write(row_index, column_index, "" if value is None else value, style)

    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="inventory_adjustments.xls",
        mimetype="application/vnd.ms-excel",
    )


def calculate_stock_movement(item, data):
    movement = data.get("movement")
    if not isinstance(movement, dict):
        packaging = infer_item_packaging(item)
        if (
            item.get("category") in CHEMICAL_CATEGORIES
            or item.get("category") == RUBBER_BLANKET_CATEGORY
            or item.get("category") == ROLL_PAPER_CATEGORY
            or is_roll_unit(item.get("unit"))
        ):
            quantity_change, quantity_error = parse_number(
                data.get("quantity_change"), "quantity_change", allow_negative=True
            )
        else:
            quantity_change, quantity_error = parse_integer(
                data.get("quantity_change"), "quantity_change", allow_negative=True
            )
        if not quantity_error and packaging and packaging["kind"] == "chemical":
            container_change = abs(quantity_change) / packaging["pack_size"]
            if not quantities_match(container_change, round(container_change)):
                return None, "chemical stock movement must equal a whole number of containers", None
        return quantity_change, quantity_error, {"quantity_change": quantity_change}

    direction = clean_text(movement.get("direction"))
    if direction not in {"in", "out"}:
        return None, "movement direction must be in or out", None
    packaging = infer_item_packaging(item)
    if not packaging:
        return None, "packaging movement is not supported for this item", None

    boxes, boxes_error = parse_optional_integer(movement.get("boxes"), "boxes")
    loose_units, loose_error = parse_optional_integer(movement.get("loose_units"), "loose units")
    if boxes_error or loose_error:
        return None, boxes_error or loose_error, None
    boxes = boxes or 0
    loose_units = loose_units or 0

    kind = packaging["kind"]
    if kind == "rubber_blankets":
        storage_type = item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE
        if storage_type == M3Z_CUT_PIECE_STORAGE_TYPE:
            sheets, sheets_error = parse_positive_integer(movement.get("sheets"), "sheets")
            if sheets_error:
                return None, sheets_error, None
            quantity_change = sheets if direction == "in" else -sheets
            return quantity_change, None, {
                "direction": direction,
                "mode": "sheets",
                "sheets": sheets,
                "normalized_quantity_change": quantity_change,
                "normalized_unit": M3Z_SHEET_STOCK_UNIT,
                "packaging": packaging,
            }
        if parse_optional_text(movement.get("rolls")) is not None:
            rolls, rolls_error = parse_positive_integer(movement.get("rolls"), "rolls")
            if rolls_error:
                return None, rolls_error, None
            breakdown = build_stock_breakdown(item) or {}
            area_per_roll = breakdown.get("area_per_roll_sqm")
            if not area_per_roll or area_per_roll <= 0:
                return None, "blanket actual width and length must be valid for stock movement", None
            quantity_change = area_per_roll * rolls
            if direction == "out":
                quantity_change = -quantity_change
            return quantity_change, None, {
                "direction": direction,
                "mode": "rolls",
                "rolls": rolls,
                "area_per_roll_sqm": area_per_roll,
                "normalized_quantity_change": quantity_change,
                "normalized_unit": RUBBER_BLANKET_STOCK_UNIT,
                "packaging": packaging,
            }

        area, area_error = parse_positive_number(
            movement.get("quantity", movement.get("area")), "square-metre movement"
        )
        if area_error:
            return None, area_error, None
        supplied_unit = movement.get("unit") or RUBBER_BLANKET_STOCK_UNIT
        if not normalize_roll_stock_unit(supplied_unit):
            return None, "partial Rubber Blanket movement unit must be m²", None
        quantity_change = area if direction == "in" else -area
        return quantity_change, None, {
            "direction": direction,
            "mode": "partial_area",
            "normalized_quantity_change": quantity_change,
            "normalized_unit": RUBBER_BLANKET_STOCK_UNIT,
            "packaging": packaging,
        }

    if kind == "calibrated_underpacking_paper":
        storage_type = item.get("storage_type") or M3Z_ROLL_STORAGE_TYPE
        if storage_type == M3Z_CUT_PIECE_STORAGE_TYPE:
            sheets, sheets_error = parse_positive_integer(movement.get("sheets"), "sheets")
            if sheets_error:
                return None, sheets_error, None
            quantity_change = sheets if direction == "in" else -sheets
            return quantity_change, None, {
                "direction": direction,
                "mode": "sheets",
                "sheets": sheets,
                "normalized_quantity_change": quantity_change,
                "normalized_unit": M3Z_SHEET_STOCK_UNIT,
                "packaging": packaging,
            }
        rolls, rolls_error = parse_positive_integer(movement.get("rolls"), "rolls")
        if rolls_error:
            return None, rolls_error, None
        breakdown = build_stock_breakdown(item) or {}
        area_per_roll = breakdown.get("area_per_roll_sqm")
        if not area_per_roll or area_per_roll <= 0:
            return None, "roll width and length must be valid for stock movement", None
        quantity_change = area_per_roll * rolls
        if direction == "out":
            quantity_change = -quantity_change
        return quantity_change, None, {
            "direction": direction,
            "rolls": rolls,
            "area_per_roll_sqm": area_per_roll,
            "normalized_quantity_change": quantity_change,
            "normalized_unit": ROLL_PAPER_STOCK_UNIT,
            "packaging": packaging,
        }
    if kind == "creasing_matrix":
        if loose_units >= packaging["units_per_box"]:
            return None, f"loose packets must be less than {packaging['units_per_box']}", None
        normalized_amount = boxes * packaging["units_per_box"] + loose_units
    elif kind == "chemical":
        if loose_units >= packaging["containers_per_box"]:
            return None, f"loose containers must be less than {packaging['containers_per_box']}", None
        containers = boxes * packaging["containers_per_box"] + loose_units
        normalized_amount = round(containers * packaging["pack_size"], 4)
    elif kind == "ctcp_plates":
        if loose_units:
            return None, "CTCP Plates stock movement only accepts full boxes", None
        normalized_amount = boxes
    else:
        return None, "unsupported packaging movement", None

    if normalized_amount <= 0:
        return None, "stock movement must be greater than 0", None
    quantity_change = normalized_amount if direction == "in" else -normalized_amount
    return quantity_change, None, {
        "direction": direction,
        "boxes": boxes,
        "loose_units": loose_units,
        "normalized_quantity_change": quantity_change,
        "normalized_unit": item.get("unit"),
        "packaging": packaging,
    }


def get_changed_fields(existing_item, item):
    comparable_fields = [
        "quantity",
        "unit",
        "batch_roll_no",
        "size",
        "width",
        "height",
        "thickness",
        "packaging",
        "format_size",
        "format_unit",
        "width_unit",
        "length",
        "length_unit",
        "thickness_unit",
        "thickness_micron",
        "storage_type",
        "number_of_rolls",
        "number_of_sheets",
        "width_meters",
        "length_meters",
        "area_per_roll_sqm",
        "area_per_sheet_sqm",
        "blanket_name",
        "nominal_width",
        "actual_width",
        "actual_width_meters",
        "roll_no",
        "batch_no",
        "print_type",
    ]
    changes = {}
    for field in comparable_fields:
        before = existing_item.get(field)
        after = item.get(field)
        if before != after:
            changes[field] = {"before": before, "after": after}
    return changes


@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/<path:filename>")
def frontend_assets(filename):
    if filename in {"style.css", "script.js", "warehouse.css", "warehouse.js"}:
        return send_from_directory(FRONTEND_DIR, filename)
    if filename == "data/inventory-config.json":
        return send_from_directory(FRONTEND_DIR / "data", "inventory-config.json")
    if filename == "favicon.ico":
        return ("", 204)
    return ("Not Found", 404)


@app.route("/health")
def health():
    try:
        database = get_database()
        database.command("ping")
        return {
            "status": "ok",
            "database": database.name,
        }
    except Exception:
        return (
            {
                "status": "degraded",
                "database": "unavailable",
                "error": "Database connection unavailable",
            },
            503,
        )


@app.errorhandler(PyMongoError)
def handle_database_error(_error):
    return jsonify({"error": "Database connection unavailable"}), 503


@app.route("/auth/me", methods=["GET"])
def auth_me():
    user, error_response = require_auth()
    if error_response:
        return jsonify({"user": None}), 200
    return jsonify({"user": serialize_user(user)})


@app.route("/auth/signup", methods=["POST"])
def auth_signup():
    users_collection = get_users_collection()
    data = request.get_json(silent=True) or {}
    email = normalize_email(data.get("email"))
    password = data.get("password")
    if not email:
        return jsonify({"error": "Email is required"}), 400

    email_error = validate_signup_email(email)
    if email_error:
        return jsonify({"error": email_error}), 400

    password_error = validate_password(password)
    if password_error:
        return jsonify({"error": password_error}), 400

    role = "user"
    if email == SUPERADMIN_EMAIL:
        role = "admin"

    now = now_ist()
    user = {
        "email": email,
        "password_hash": generate_password_hash(password.strip()),
        "role": role,
        "created_at": now,
        "updated_at": now,
    }

    try:
        result = users_collection.insert_one(user)
    except DuplicateKeyError:
        return jsonify({"error": "An account with this email already exists"}), 409

    user["_id"] = result.inserted_id
    session["user_id"] = str(result.inserted_id)
    return jsonify({"message": "Signup successful", "user": serialize_user(user)}), 201


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(silent=True) or {}
    email = normalize_email(data.get("email"))
    password = data.get("password")

    if not email or not isinstance(password, str):
        return jsonify({"error": "Email and password are required"}), 400

    if email != SUPERADMIN_EMAIL and not email.endswith(ALLOWED_SIGNUP_DOMAIN):
        return jsonify({"error": "Not permitted to login"}), 403

    user = get_users_collection().find_one({"email": email})
    if not user and email == SUPERADMIN_EMAIL:
        password_error = validate_password(password)
        if password_error:
            return jsonify({"error": password_error}), 400
        user, _ = create_superadmin_if_missing(password)

    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "Invalid email or password"}), 401

    session["user_id"] = str(user["_id"])
    return jsonify({"message": "Login successful", "user": serialize_user(user)})


@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"message": "Logged out"})


@app.route("/auth/forgot-password", methods=["POST"])
def auth_forgot_password():
    data = request.get_json(silent=True) or {}
    email = normalize_email(data.get("email"))
    new_password = data.get("new_password")

    if not email:
        return jsonify({"error": "Email is required"}), 400

    password_error = validate_password(new_password)
    if password_error:
        return jsonify({"error": password_error}), 400

    user = get_users_collection().find_one({"email": email})
    if not user:
        return jsonify({"error": "Account not found"}), 404

    get_users_collection().update_one(
        {"_id": user["_id"]},
        {
            "$set": {
                "password_hash": generate_password_hash(new_password.strip()),
                "updated_at": now_ist(),
            }
        },
    )
    session["user_id"] = str(user["_id"])
    updated_user = get_users_collection().find_one({"_id": user["_id"]})
    return jsonify({"message": "Password reset successful", "user": serialize_user(updated_user)})


@app.route("/add-item", methods=["POST"])
def add_item():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    data = request.get_json(silent=True) or {}
    item, error = build_item_payload(data)
    reason = get_request_reason(data, "Manual item creation")

    if error:
        return jsonify({"error": error}), 400

    try:
        result = inventory_collection.insert_one(item)
    except DuplicateKeyError:
        return jsonify({"error": "Item already exists"}), 409

    item["_id"] = result.inserted_id
    log_stock_change(item, "created", 0, item["quantity"], "manual", reason)
    return jsonify({"message": "Item added", "item": serialize_item(item)}), 201


@app.route("/inventory", methods=["GET"])
def get_inventory():
    _, error_response = require_role("user", "workshop", "admin")
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    query, error = create_inventory_query(request.args)
    if error:
        return jsonify({"error": error}), 400

    items = inventory_collection.find(query).sort(get_inventory_sort())
    return jsonify([serialize_item(item) for item in items])


@app.route("/stock-logs", methods=["GET"])
def get_stock_logs():
    _, error_response = require_role(*LOG_ACCESS_ROLES)
    if error_response:
        return error_response
    stock_logs_collection = get_stock_logs_collection()
    limit, error = parse_integer(request.args.get("limit", 50), "limit")
    if error:
        return jsonify({"error": error}), 400

    logs = stock_logs_collection.find().sort("changed_at", -1).limit(limit)
    return jsonify([serialize_log(log) for log in logs])


@app.route("/admin/users", methods=["GET"])
def get_users():
    _, error_response = require_role("admin")
    if error_response:
        return error_response

    users = get_users_collection().find().sort("created_at", 1)
    return jsonify([serialize_user(user) for user in users])


@app.route("/admin/users/<user_id>/role", methods=["PUT"])
def update_user_role(user_id):
    current_user, error_response = require_role("admin")
    if error_response:
        return error_response

    data = request.get_json(silent=True) or {}
    role = clean_text(data.get("role"))
    if role not in ALLOWED_ROLES:
        return jsonify({"error": "Role must be user, admin, or workshop"}), 400

    try:
        target_object_id = ObjectId(user_id)
    except Exception:
        return jsonify({"error": "Invalid user id"}), 400

    users_collection = get_users_collection()
    target_user = users_collection.find_one({"_id": target_object_id})
    if not target_user:
        return jsonify({"error": "User not found"}), 404

    if target_user["email"] == SUPERADMIN_EMAIL and role != "admin":
        return jsonify({"error": "Superadmin role cannot be changed"}), 400

    users_collection.update_one(
        {"_id": target_object_id},
        {"$set": {"role": role, "updated_at": now_ist()}},
    )
    updated_user = users_collection.find_one({"_id": target_object_id})
    return jsonify({"message": "User role updated", "user": serialize_user(updated_user)})


@app.route("/admin/users/<user_id>", methods=["DELETE"])
def delete_user(user_id):
    current_user, error_response = require_role("admin")
    if error_response:
        return error_response

    try:
        target_object_id = ObjectId(user_id)
    except Exception:
        return jsonify({"error": "Invalid user id"}), 400

    users_collection = get_users_collection()
    target_user = users_collection.find_one({"_id": target_object_id})
    if not target_user:
        return jsonify({"error": "User not found"}), 404

    if target_user["email"] == SUPERADMIN_EMAIL:
        return jsonify({"error": "Superadmin cannot be removed"}), 400

    if str(target_user["_id"]) == str(current_user["_id"]):
        return jsonify({"error": "You cannot remove your own account"}), 400

    users_collection.delete_one({"_id": target_object_id})
    return jsonify({"message": "User removed"})


@app.route("/update-stock", methods=["PUT"])
def update_stock():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    data = request.get_json(silent=True) or {}
    lookup, error = build_lookup(data)
    reason = get_request_reason(data, "Manual stock movement")
    if error:
        return jsonify({"error": error}), 400

    item = inventory_collection.find_one(lookup)
    if not item:
        return jsonify({"error": "Item not found"}), 404

    quantity_change, quantity_error, movement_details = calculate_stock_movement(item, data)
    if quantity_error:
        return jsonify({"error": quantity_error}), 400

    new_quantity = item["quantity"] + quantity_change
    if new_quantity < -0.000001:
        return jsonify({"error": "Quantity cannot go below 0"}), 400
    if abs(new_quantity) < 0.000001:
        new_quantity = 0

    updates = {
        "quantity": new_quantity,
        "updated_at": now_ist(),
    }
    item_packaging = infer_item_packaging(item)
    if item_packaging and item_packaging["kind"] in {"calibrated_underpacking_paper", "rubber_blankets"}:
        movement_rolls = (movement_details or {}).get("rolls")
        movement_sheets = (movement_details or {}).get("sheets")
        if movement_sheets is not None:
            current_sheets = item.get("number_of_sheets")
            if current_sheets is None:
                current_sheets = int(round(item.get("quantity", 0) or 0))
            else:
                current_sheets, current_sheets_error = parse_integer(current_sheets, "number of sheets")
                if current_sheets_error:
                    return jsonify({"error": "stored sheet count is invalid"}), 400
            next_sheets = current_sheets + (
                movement_sheets if movement_details.get("direction") == "in" else -movement_sheets
            )
            if next_sheets < 0:
                return jsonify({"error": "number of sheets cannot go below 0"}), 400
            updates["number_of_sheets"] = next_sheets
        elif movement_rolls is not None:
            current_rolls = item.get("number_of_rolls")
            if not isinstance(current_rolls, int):
                current_rolls, current_rolls_error = parse_integer(current_rolls, "number of rolls")
                if current_rolls_error:
                    return jsonify({"error": "stored roll count is invalid"}), 400
            next_rolls = current_rolls + (movement_rolls if movement_details.get("direction") == "in" else -movement_rolls)
            if next_rolls < 0:
                return jsonify({"error": "number of rolls cannot go below 0"}), 400
            if item_packaging["kind"] == "rubber_blankets" and (item.get("roll_no") or item.get("batch_no")) and next_rolls > 1:
                return jsonify({"error": "identified blanket rolls must be received as separate inventory rows"}), 400
            updates["number_of_rolls"] = next_rolls
        elif item_packaging["kind"] == "rubber_blankets" and new_quantity == 0:
            if item.get("storage_type") == M3Z_CUT_PIECE_STORAGE_TYPE:
                updates["number_of_sheets"] = 0
            else:
                updates["number_of_rolls"] = 0
    inferred_packaging = item_packaging
    if inferred_packaging:
        updates["packaging"] = inferred_packaging

    if data.get("thickness") is not None:
        updates["thickness"] = parse_optional_text(data.get("thickness"))

    if data.get("unit") is not None:
        unit = clean_text(data.get("unit"))
        if not unit:
            return jsonify({"error": "unit cannot be empty"}), 400
        packaging = infer_item_packaging(item)
        expected_unit = None
        if packaging:
            expected_unit = {
                "creasing_matrix": "pkt",
                "ctcp_plates": "box",
                "chemical": packaging.get("pack_unit"),
                "rubber_blankets": RUBBER_BLANKET_STOCK_UNIT,
                "calibrated_underpacking_paper": ROLL_PAPER_STOCK_UNIT,
            }.get(packaging["kind"])
        normalized_unit = "ltr" if unit.lower() == "l" else unit.lower()
        if packaging and packaging["kind"] in {"rubber_blankets", "calibrated_underpacking_paper"}:
            normalized_unit = normalize_roll_stock_unit(unit)
        if packaging and packaging["kind"] in {"rubber_blankets", "calibrated_underpacking_paper"} and item.get("storage_type") == M3Z_CUT_PIECE_STORAGE_TYPE:
            expected_unit = M3Z_SHEET_STOCK_UNIT
            normalized_unit = normalize_sheet_stock_unit(unit)
        if expected_unit and normalized_unit != expected_unit:
            return jsonify({"error": f"unit must remain {expected_unit} for this item"}), 400
        unit = expected_unit or unit
        updates["unit"] = unit

    updates["size"] = build_size_label(item.get("width"), item.get("height"))

    inventory_collection.update_one(lookup, {"$set": updates})
    updated_item = inventory_collection.find_one(lookup)
    log_item = dict(item)
    log_item["unit"] = updated_item["unit"]
    log_stock_change(
        log_item,
        "updated",
        item["quantity"],
        updated_item["quantity"],
        "manual",
        reason,
        {
            "movement": movement_details,
            "quantity_before": item["quantity"],
            "quantity_after": updated_item["quantity"],
        },
    )

    return jsonify({"message": "Stock updated", "item": serialize_item(updated_item)})


@app.route("/delete-item", methods=["DELETE"])
def delete_item():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    data = request.get_json(silent=True) or {}
    lookup, error = build_lookup(data)
    reason = get_request_reason(data, "Manual item deletion")
    if error:
        return jsonify({"error": error}), 400

    item = inventory_collection.find_one(lookup)
    if not item:
        return jsonify({"error": "Item not found"}), 404

    inventory_collection.delete_one({"_id": item["_id"]})
    log_stock_change(item, "deleted", item["quantity"], 0, "manual", reason)
    return jsonify({"message": "Item deleted"})


@app.route("/upload-excel", methods=["POST"])
def upload_excel():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    uploaded_file = request.files.get("file")
    upload_mode = (request.form.get("mode") or "import").strip().lower()
    action = (request.form.get("action") or "preview").strip().lower()
    if upload_mode not in {"import", "update"}:
        return jsonify({"error": "mode must be import or update"}), 400
    if action not in {"preview", "apply"}:
        return jsonify({"error": "action must be preview or apply"}), 400

    if uploaded_file is None or uploaded_file.filename == "":
        return jsonify({"error": "Excel file is required"}), 400

    parsed_workbook, parse_error = parse_uploaded_excel_workbook(uploaded_file)
    if parse_error:
        return jsonify(
            {
                "error": parse_error,
                "expected_sheets": [definition["sheet_name"] for definition in EXCEL_CATEGORY_SHEETS],
            }
        ), 400

    preview = build_excel_preview_response(parsed_workbook, inventory_collection, upload_mode)
    if action == "preview":
        return jsonify(preview)

    if not preview["can_apply"]:
        return jsonify({"error": "Correct all workbook errors before applying inventory changes", "preview": preview}), 400
    if upload_mode == "update" and str(request.form.get("confirm_update", "")).lower() not in {"1", "true", "yes"}:
        return jsonify({"error": "Update confirmation is required before records can be deleted"}), 400
    reason = get_excel_reason()
    result = apply_parsed_excel_workbook(parsed_workbook, inventory_collection, upload_mode, reason)
    return jsonify(result)


@app.route("/download-import-template", methods=["GET"])
def download_import_template():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    return send_inventory_workbook([], "only_stock_import_template.xlsx", "template")


@app.route("/export-update-excel", methods=["GET"])
def export_update_excel():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    items = list(inventory_collection.find().sort(get_inventory_sort()))
    return send_inventory_workbook(items, "only_stock_update_workbook.xlsx", "update")


@app.route("/export-excel", methods=["GET"])
def export_excel():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response
    inventory_collection = get_inventory_collection()
    items = list(inventory_collection.find().sort(get_inventory_sort()))
    return send_inventory_workbook(items, "only_stock_current_inventory.xlsx", "current")


@app.route("/inventory-adjustments/export", methods=["POST"])
def export_inventory_adjustments():
    _, error_response = require_role(*WRITE_ROLES)
    if error_response:
        return error_response

    pasted_text = request.form.get("items", "")
    uploaded_file = request.files.get("file")
    file_text = ""
    if uploaded_file and uploaded_file.filename:
        if not uploaded_file.filename.lower().endswith(".xls"):
            return jsonify({"error": "Inventory adjustment upload must be an .xls file"}), 400
        try:
            file_text = extract_adjustment_text_from_xls(uploaded_file)
        except Exception:
            return jsonify({"error": "Unable to read inventory adjustment .xls file"}), 400

    source_text = "\n".join(part for part in [file_text, pasted_text] if part.strip())
    if not source_text.strip():
        return jsonify({"error": "Add item text or upload an .xls file"}), 400

    defaults = {
        "date_display": normalized_adjustment_date(request.form.get("date")),
        "reason": FIXED_ADJUSTMENT_REASON,
        "unit": "Mtr",
    }
    rows = parse_inventory_adjustment_text(source_text, defaults)
    if not rows:
        return jsonify({"error": "No adjustment rows found. Use: Product Name, then R13 - 27.50 Mtr - 15.55 - 2559.04"}), 400

    return send_inventory_adjustment_xls(rows)


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(debug=debug, host="0.0.0.0", port=port)
