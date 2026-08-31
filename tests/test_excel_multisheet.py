import unittest
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from backend.app import (
    EXCEL_CATEGORY_SHEETS,
    EXCEL_LISTS_SHEET,
    EXCEL_README_SHEET,
    apply_parsed_excel_workbook,
    build_inventory_workbook,
    build_item_key,
    build_item_payload,
    find_excel_update_deletions,
    get_excel_update_scopes,
    get_parsed_excel_items,
    parse_uploaded_excel_workbook,
    app,
)


class ExcelUpload(BytesIO):
    filename = "inventory.xlsx"


def matches_query(item, query):
    for field, expected in query.items():
        actual = item.get(field)
        if isinstance(expected, dict):
            if "$in" in expected and actual not in expected["$in"]:
                return False
            if "$exists" in expected and (field in item) != expected["$exists"]:
                return False
            if "$ne" in expected and actual == expected["$ne"]:
                return False
        elif actual != expected:
            return False
    return True


class ReadOnlyCollection:
    def __init__(self, items=None):
        self.items = list(items or [])

    def find(self, query=None):
        query = query or {}
        return [item for item in self.items if matches_query(item, query)]

    def find_one(self, query):
        return next((item for item in self.items if matches_query(item, query)), None)


class MutableCollection(ReadOnlyCollection):
    def __init__(self, items=None):
        super().__init__(items)
        self.insert_calls = 0

    def insert_one(self, item):
        self.insert_calls += 1
        item["_id"] = 1000 + self.insert_calls
        self.items.append(item)
        return SimpleNamespace(inserted_id=item["_id"])

    def update_one(self, query, update):
        item = self.find_one(query)
        if item:
            item.update(update.get("$set", {}))

    def delete_one(self, query):
        item = self.find_one(query)
        if item:
            self.items.remove(item)


class MultiSheetExcelTests(unittest.TestCase):
    def build_item(self, payload):
        item, error = build_item_payload(payload)
        self.assertIsNone(error)
        return item

    def workbook_with_rows(self, rows_by_sheet):
        payload = build_inventory_workbook([], "template")
        workbook = load_workbook(BytesIO(payload.getvalue()))
        for sheet_name, rows in rows_by_sheet.items():
            for row_index, values in enumerate(rows, start=5):
                for column_index, value in enumerate(values, start=1):
                    workbook[sheet_name].cell(row_index, column_index, value)
        output = BytesIO()
        workbook.save(output)
        return ExcelUpload(output.getvalue())

    def test_template_has_professional_multi_sheet_structure(self):
        payload = build_inventory_workbook([], "template")
        workbook = load_workbook(BytesIO(payload.getvalue()))
        expected = [EXCEL_README_SHEET, *[definition["sheet_name"] for definition in EXCEL_CATEGORY_SHEETS], EXCEL_LISTS_SHEET]
        self.assertEqual(workbook.sheetnames, expected)
        self.assertEqual(workbook.active.title, EXCEL_README_SHEET)
        self.assertEqual(workbook[EXCEL_LISTS_SHEET].sheet_state, "veryHidden")
        self.assertTrue(all(len(name) <= 31 for name in workbook.sheetnames))
        for definition in EXCEL_CATEGORY_SHEETS:
            headers = [cell.value for cell in workbook[definition["sheet_name"]][4]]
            self.assertEqual(headers, definition["headers"])
            self.assertEqual(workbook[definition["sheet_name"]].freeze_panes, "A5")

    def test_all_category_sheets_recalculate_backend_quantities(self):
        upload = self.workbook_with_rows(
            {
                "01_Rubber_Blankets": [
                    ["Magnum SF", "Roll", 1.95, "mm", None, 780, 790, "mm", 30, "m", 5, None],
                    ["Magnum SF", "Cut Piece", 1.95, "mm", None, 780, 790, "mm", 795, "mm", None, 50],
                ],
                "02_Calibrated_Underpacking": [
                    ["Roll", 400, 1320, "mm", 100, "m", 8, None],
                    ["Cut Piece", 400, 1320, "mm", 795, "mm", None, 50],
                ],
                "03_Creasing_Matrix": [[11, "0.8 X 2.5", 2, 3, 23]],
                "04_Chemical": [["Chem R-ol", 5, "bottle", 4, 2, 3, 11, 55, "L"]],
                "05_CTCP_Plates": [[0.30, "650 X 550", 7, 50, 350, 7, "Box"]],
            }
        )
        parsed, error = parse_uploaded_excel_workbook(upload)
        self.assertIsNone(error)
        items = get_parsed_excel_items(parsed)
        self.assertEqual(len(items), 7)
        quantities = [(item["category"], item.get("storage_type"), item["quantity"], item["unit"]) for item in items]
        self.assertIn(("Rubber Blankets", "roll", 118.5, "m²"), quantities)
        self.assertIn(("Rubber Blankets", "cut_piece", 50, "sheets"), quantities)
        self.assertIn(("Calibrated Underpacking Paper", "roll", 1056.0, "m²"), quantities)
        self.assertIn(("Calibrated Underpacking Paper", "cut_piece", 50, "sheets"), quantities)
        self.assertIn(("Creasing Matrix", None, 23, "pkt"), quantities)
        self.assertIn(("Washing Solutions", None, 55, "ltr"), quantities)
        self.assertIn(("CTCP Plates", None, 7, "box"), quantities)

    def test_legacy_single_sheet_workbook_is_still_detected(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Import Items"
        headers = ["Category", "Brand", "Type", "Width", "Length", "Thickness", "Quantity", "Unit"]
        sheet.append(headers)
        sheet.append(["Spray Powder", "Generic", "Standard", None, None, None, 8, "pcs"])
        output = BytesIO()
        workbook.save(output)
        parsed, error = parse_uploaded_excel_workbook(ExcelUpload(output.getvalue()))
        self.assertIsNone(error)
        self.assertEqual(parsed["workbook_type"], "legacy")
        self.assertEqual(len(get_parsed_excel_items(parsed)), 1)

    def test_update_deletion_is_scoped_to_non_empty_provided_sheet(self):
        upload = self.workbook_with_rows(
            {
                "02_Calibrated_Underpacking": [["Roll", 400, 1320, "mm", 100, "m", 8, None]],
            }
        )
        parsed, error = parse_uploaded_excel_workbook(upload)
        self.assertIsNone(error)
        kept_m3z = get_parsed_excel_items(parsed)[0]
        missing_m3z = self.build_item(
            {
                "category": "Calibrated Underpacking Paper",
                "storage_type": "roll",
                "thickness_micron": 200,
                "width": 1240,
                "width_unit": "mm",
                "length": 100,
                "length_unit": "m",
                "number_of_rolls": 2,
            }
        )
        chemical = self.build_item(
            {
                "category": "Washing Solutions",
                "product": "Chem R-ol",
                "containers_per_box": 4,
                "boxes": 1,
                "loose_units": 0,
                "unit": "ltr",
            }
        )
        for identifier, item in enumerate([kept_m3z, missing_m3z, chemical], start=1):
            item["_id"] = identifier
        collection = ReadOnlyCollection([kept_m3z, missing_m3z, chemical])
        scopes = get_excel_update_scopes(parsed)
        self.assertEqual([scope["key"] for scope in scopes], ["calibrated_underpacking_paper"])
        deletions = find_excel_update_deletions(parsed, collection, {build_item_key(kept_m3z)})
        self.assertEqual([item["_id"] for item in deletions], [2])

    def test_upload_route_previews_before_any_inventory_mutation(self):
        workbook = self.workbook_with_rows(
            {"02_Calibrated_Underpacking": [["Roll", 400, 1320, "mm", 100, "m", 8, None]]}
        ).getvalue()
        collection = MutableCollection()
        client = app.test_client()
        with (
            patch("backend.app.require_role", return_value=({"role": "admin"}, None)),
            patch("backend.app.get_inventory_collection", return_value=collection),
            patch("backend.app.log_stock_change"),
        ):
            preview_response = client.post(
                "/upload-excel",
                data={
                    "file": (BytesIO(workbook), "inventory.xlsx"),
                    "mode": "import",
                    "action": "preview",
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(preview_response.status_code, 200)
            self.assertTrue(preview_response.get_json()["can_apply"])
            self.assertEqual(collection.insert_calls, 0)

            apply_response = client.post(
                "/upload-excel",
                data={
                    "file": (BytesIO(workbook), "inventory.xlsx"),
                    "mode": "import",
                    "action": "apply",
                    "reason": "Test import",
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(apply_response.status_code, 200)
            self.assertEqual(apply_response.get_json()["inserted"], 1)
            self.assertEqual(collection.insert_calls, 1)

            unconfirmed_update = client.post(
                "/upload-excel",
                data={
                    "file": (BytesIO(workbook), "inventory.xlsx"),
                    "mode": "update",
                    "action": "apply",
                    "reason": "Unconfirmed test update",
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(unconfirmed_update.status_code, 400)
            self.assertIn("confirmation", unconfirmed_update.get_json()["error"].lower())

    def test_confirmed_update_deletes_only_records_in_non_empty_scope(self):
        upload = self.workbook_with_rows(
            {"02_Calibrated_Underpacking": [["Roll", 400, 1320, "mm", 100, "m", 8, None]]}
        )
        parsed, error = parse_uploaded_excel_workbook(upload)
        self.assertIsNone(error)
        kept_m3z = get_parsed_excel_items(parsed)[0]
        missing_m3z = self.build_item(
            {
                "category": "Calibrated Underpacking Paper",
                "storage_type": "roll",
                "thickness_micron": 200,
                "width": 1240,
                "width_unit": "mm",
                "length": 100,
                "length_unit": "m",
                "number_of_rolls": 2,
            }
        )
        chemical = self.build_item(
            {
                "category": "Washing Solutions",
                "product": "Chem R-ol",
                "containers_per_box": 4,
                "boxes": 1,
                "loose_units": 0,
                "unit": "ltr",
            }
        )
        for identifier, item in enumerate([kept_m3z, missing_m3z, chemical], start=1):
            item["_id"] = identifier
        collection = MutableCollection([kept_m3z, missing_m3z, chemical])
        with patch("backend.app.log_stock_change") as log_change:
            result = apply_parsed_excel_workbook(parsed, collection, "update", "Scoped update test")
        self.assertEqual(result["deleted"], 1)
        self.assertEqual([item["_id"] for item in collection.items], [1, 3])
        self.assertTrue(any(call.args[1] == "excel_delete" for call in log_change.call_args_list))


if __name__ == "__main__":
    unittest.main()
